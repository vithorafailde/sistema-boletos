import json, re, base64, io, unicodedata, time, os, traceback
import urllib.request
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import urllib.request
import urllib.error
import json as _json
from pathlib import Path
from functools import wraps
from datetime import datetime, date, timedelta
from flask import Flask, render_template, request, jsonify, Response, stream_with_context, session, redirect, url_for
import pdfplumber
from openpyxl import load_workbook
import anthropic

BASE = Path(__file__).parent
UPLOAD_DIR = BASE / "uploads"
DATA_DIR = BASE / "data"
CONFIG_FILE = DATA_DIR / "config.json"
HISTORICO_FILE = DATA_DIR / "historico.json"

DIMOB_HISTORICO_FILE = DATA_DIR / "dimob_historico.json"
LOCATARIOS_EMAILS_FILE = DATA_DIR / "locatarios_emails.json"
LOG_ENVIOS_FILE = DATA_DIR / "log_envios.json"

MESES_NOMES = ['Janeiro','Fevereiro','Março','Abril','Maio','Junho',
               'Julho','Agosto','Setembro','Outubro','Novembro','Dezembro']

UPLOAD_DIR.mkdir(exist_ok=True)
DATA_DIR.mkdir(exist_ok=True)

# Tudo que vier no boleto da unidade é repassado ao locatário,
# exceto correio e fundo_reserva que são encargos do proprietário.
REPASSE_ITENS = ["agua", "gas", "energia", "tx_leitura", "vaga_moto", "tag"]
# Encargos do PROPRIETÁRIO: não repassar ao locatário
NAO_REPASSE = ["fundo_reserva", "correio", "melhorias", "outros"]

# Edifícios onde NÃO se repassa NENHUM custo de condomínio ao locatário (só aluguel)
EDIFICIOS_EXCLUIDOS = ["lamelas", "paisagem", "victoria"]

# Palavras-chave no nome/endereço do edifício que indicam consumo individual (cv / vm)
# Usadas em montar_resultado() para determinar se o boleto tem água/gás medidos.
# Complementam o regex \b(cv|vm)\b aplicado ao nome do arquivo.
EDIFICIOS_CONSUMO_KEYWORDS = [
    "casa verde", "mandaqui", "condominio cv",
    "cond. cv", "cond cv", "vila jardim cv",
]

# Séries de variação mensal (%) no SGS/BACEN — usadas para contratos FUTURO (dados parciais)
# IGPM usa série 28655 (alta precisão, ~10 decimais) em vez da 189 (2 decimais)
BACEN_SERIES = {'IPCA': 433, 'IGPM': 28655, 'INPC': 188}
# Séries de número-índice no SGS/BACEN — usadas para contratos ESTE_MES (precisão máxima)
BACEN_SERIES_INDICE = {'IPCA': 1737, 'INPC': 1617}

LABELS = {
    "cota": "Cota Cond.", "agua": "Cons. Agua", "gas": "Cons. Gas",
    "energia": "Energia Comum", "tx_leitura": "Tx. Leitura",
    "vaga_moto": "Vaga/Garagem", "tag": "Tag/Acesso", "melhorias": "Melhorias",
    "outros": "Outros Cond.", "fundo_reserva": "Fundo Reserva", "correio": "Correio",
}

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "boletos_sistema_2026_xK9#mP")
app.config["MAX_CONTENT_LENGTH"] = 150 * 1024 * 1024

@app.after_request
def no_cache(response):
    if "text/html" in response.content_type:
        response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        response.headers["Pragma"] = "no-cache"
        response.headers["Expires"] = "0"
    return response

# ─── Credenciais de acesso ────────────────────────────────────────────────────
LOGIN_USUARIO = os.environ.get("LOGIN_USUARIO", "vitho")
LOGIN_SENHA   = os.environ.get("LOGIN_SENHA",   "vi28041305")
LOGIN_USUARIO_2 = os.environ.get("LOGIN_USUARIO_2", "lourdes")
LOGIN_SENHA_2   = os.environ.get("LOGIN_SENHA_2",   "vi280407")

USUARIOS = {
    LOGIN_USUARIO:   "admin",
    LOGIN_USUARIO_2: "user",
}
SENHAS = {
    LOGIN_USUARIO:   LOGIN_SENHA,
    LOGIN_USUARIO_2: LOGIN_SENHA_2,
}

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("logado"):
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("logado"):
            return redirect(url_for("login"))
        if session.get("role") != "admin":
            return jsonify({"ok": False, "erro": "Acesso restrito ao administrador."}), 403
        return f(*args, **kwargs)
    return decorated

def is_admin():
    return session.get("role") == "admin"


# ─── Normalização ──────────────────────────────────────────────────────────────

def norm(nome):
    if not nome:
        return ""
    n = unicodedata.normalize("NFKD", str(nome).upper())
    n = "".join(c for c in n if not unicodedata.combining(c))
    return re.sub(r"[^A-Z0-9]", "", n)

def norm_palavras(nome):
    if not nome:
        return []
    n = unicodedata.normalize("NFKD", str(nome).upper())
    n = "".join(c for c in n if not unicodedata.combining(c))
    return [w for w in re.sub(r"[^A-Z0-9 ]", " ", n).split() if len(w) > 2]

def extrair_nome_arquivo(nome_arquivo):
    """Extrai nome do locatário e unidade do padrão do nome do arquivo.
    Suporta:
      'cond_condominio vm - helio - apto 56a.pdf'      → ('helio', '56A')
      'condominio altavilla - joao - denis apto 37.pdf' → ('denis', '37')
    Regra: o ÚLTIMO segmento separado por ' - ' (antes do apto) é o locatário.
    """
    stem = Path(nome_arquivo).stem.lower()
    # Remove prefixo cond_condominio ou condominio
    stem = re.sub(r'^cond_condomini[oo]?\s*', '', stem).strip()
    stem = re.sub(r'^condomini[oo]\s*', '', stem).strip()
    # Remove primeiro token antes do ' - ' (tipo do edifício: vm, cv, altavilla…)
    stem = re.sub(r'^[a-z0-9]+\s*-\s*', '', stem).strip()
    # Extrai unidade: aceita "- apto XX" ou " apto XX" (com ou sem traço)
    unidade = ""
    m = re.search(r'(?:\s*-\s*|\s+)apto\s+([0-9]+[a-z]?)', stem, re.IGNORECASE)
    if m:
        unidade = m.group(1).upper()
        stem = stem[:m.start()].strip()
    # Se houver múltiplos segmentos (prop - loc), usa o ÚLTIMO como locatário
    partes = [p.strip() for p in re.split(r'\s*-\s*', stem) if p.strip()]
    nome = partes[-1] if partes else stem
    return nome.strip(), unidade

def score_nome_arquivo(palavras_arq, palavras_loc):
    """Score de similaridade entre nome do arquivo (apelido/abreviado) e nome do locatário.
    Aceita abreviações de 2 letras (ex: HP) com match exato."""
    validas = [p for p in palavras_arq if len(p) >= 2]
    if not validas or not palavras_loc:
        return 0.0
    matches = 0
    for pa in validas:
        for pl in palavras_loc:
            if pa == pl:
                matches += 1
                break
            # prefixo comum de 4+ chars cobre variações ortográficas (luciaria/luciara)
            # só para palavras com 4+ letras — abreviações exigem match exato
            if len(pa) >= 4:
                n = min(len(pa), len(pl), 5)
                if n >= 4 and pa[:n] == pl[:n]:
                    matches += 0.8
                    break
    return matches / len(validas)

def extrair_rua(texto):
    palavras = norm_palavras(texto)
    prefixos = {"RUA", "AVENIDA", "AV", "TRAVESSA", "ALAMEDA", "PRACA", "ESTRADA"}
    for p in palavras:
        if p not in prefixos:
            return p
    return palavras[0] if palavras else ""

def extrair_unidade(texto):
    t = str(texto).upper()
    for padrao in [
        r'(?:APTO?|AP|APARTAMENTO|UNIDADE|SALA|CONJ)\s*[./\-]?\s*(\d{1,4}[A-Z]?)',
        r'(?:NR?|N[°º])\s*(\d{1,4}[A-Z]?)\b',
    ]:
        m = re.search(padrao, t)
        if m:
            return re.sub(r"[^0-9A-Z]", "", m.group(1))
    return ""

def safe_float(val):
    """Converte valor monetário para float de forma segura.
    Aceita: None (→ 0.0), int, float, strings como '31,31' ou 'R$ 1.234,56'.
    Nunca lança exceção — retorna 0.0 em caso de falha.
    """
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    try:
        s = str(val).strip()
        # Remove R$, espaços normais e non-breaking spaces
        s = re.sub(r'[R$\s\xa0]', '', s)
        # Formato BR: ponto = milhar, vírgula = decimal → ex: "1.234,56"
        if ',' in s and '.' in s:
            s = s.replace('.', '').replace(',', '.')
        elif ',' in s:
            # Só vírgula → vírgula é decimal → ex: "31,31"
            s = s.replace(',', '.')
        return float(s) if s else 0.0
    except (ValueError, AttributeError):
        return 0.0


# ─── Reajuste de Aluguel ──────────────────────────────────────────────────────

def normalizar_indice(s):
    """Normaliza o texto do índice para uma das chaves de BACEN_SERIES."""
    s = (s or '').upper().strip()
    if 'IPCA' in s:
        return 'IPCA'
    if 'IGP' in s or 'IGPM' in s:
        return 'IGPM'
    if 'INPC' in s:
        return 'INPC'
    return None


def parse_data_reajuste(val):
    """Converte string ou datetime do Excel para objeto date.
    Aceita também formato DD/MM (sem ano) — nesse caso usa o ano corrente.
    """
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()

    # Formato completo com ano
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y'):
        try:
            return datetime.strptime(s[:10], fmt).date()
        except ValueError:
            continue

    # Formato DD/MM sem ano → usa o ano corrente
    try:
        partes = s[:5].split('/')
        if len(partes) == 2:
            dia, mes = int(partes[0]), int(partes[1])
            return date(date.today().year, mes, dia)
    except (ValueError, IndexError):
        pass

    return None


def buscar_historico_bacen(serie_id, anos=5, tentativas=3, timeout=30):
    """Busca histórico mensal de um índice no SGS/BACEN.
    Retorna (dict{YYYY-MM: float}, erro_str_ou_None).
    Tenta até `tentativas` vezes com `timeout` segundos cada.
    """
    hoje = date.today()
    ini = date(hoje.year - anos, 1, 1)
    url = (
        f"https://api.bcb.gov.br/dados/serie/bcdata.sgs.{serie_id}/dados"
        f"?formato=json&dataInicial={ini.strftime('%d/%m/%Y')}&dataFinal={hoje.strftime('%d/%m/%Y')}"
    )
    ultimo_erro = None
    for tentativa in range(1, tentativas + 1):
        try:
            req = urllib.request.Request(
                url,
                headers={'Accept': 'application/json', 'User-Agent': 'python-urllib/sistema-boletos'}
            )
            with urllib.request.urlopen(req, timeout=timeout) as r:
                dados = json.loads(r.read().decode('utf-8'))
            hist = {}
            for item in dados:
                try:
                    d = datetime.strptime(item['data'], '%d/%m/%Y')
                    hist[f"{d.year}-{d.month:02d}"] = float(str(item['valor']).replace(',', '.'))
                except Exception:
                    pass
            return hist, None
        except Exception as e:
            ultimo_erro = str(e)
            if tentativa < tentativas:
                time.sleep(2)   # aguarda 2 s antes de tentar novamente
    return {}, ultimo_erro


def buscar_acumulado12_bacen(serie_id, tentativas=3):
    """Busca a série de acumulado 12 meses do SGS/BACEN (ex: série 28655 = IGPM acum. 12m).
    Retorna (dict{YYYY-MM: float_percentual}, erro_str_ou_None).
    O valor para o mês M já é o acumulado dos 12 meses terminados em M — mesmo cálculo
    que a Calculadora do Cidadão usa, sem arredondamento acumulado.
    """
    return buscar_historico_bacen(serie_id, anos=3, tentativas=tentativas)


def buscar_indice_igpm_ipea(tentativas=3):
    """Busca o número-índice mensal do IGPM (FGV) via IPEA.
    Tenta HTTPS primeiro, depois HTTP como fallback.
    Retorna (dict{YYYY-MM: float}, erro_str_ou_None).
    """
    urls = [
        "https://www.ipeadata.gov.br/api/odata4/ValoresSerie(SERCODIGO='IGP12_IGPM12')?$select=VALDATA,VALVALOR",
        "http://www.ipeadata.gov.br/api/odata4/ValoresSerie(SERCODIGO='IGP12_IGPM12')?$select=VALDATA,VALVALOR",
    ]
    ultimo_erro = None
    for url in urls:
        for tentativa in range(1, tentativas + 1):
            try:
                req = urllib.request.Request(
                    url, headers={'Accept': 'application/json',
                                  'User-Agent': 'python-urllib/sistema-boletos'})
                with urllib.request.urlopen(req, timeout=30) as r:
                    dados = json.loads(r.read().decode('utf-8'))
                hist = {}
                for item in dados.get('value', []):
                    try:
                        d = date.fromisoformat(str(item['VALDATA'])[:10])
                        val = item.get('VALVALOR')
                        if val is not None:
                            hist[f"{d.year}-{d.month:02d}"] = float(val)
                    except Exception:
                        pass
                if hist:
                    return hist, None
                ultimo_erro = 'IPEA: nenhum dado retornado'
            except Exception as e:
                ultimo_erro = str(e)
                if tentativa < tentativas:
                    time.sleep(2)
    return {}, ultimo_erro


def calcular_por_numero_indice(hist_indice, data_aniversario):
    """Calcula variação pelo número-índice publicado (máxima precisão).
    Fórmula: Index(mes_aniv_atual) / Index(mês_anterior_ao_aniv_ano_anterior) − 1
    = 13 variações mensais — equivalente ao BACEN Cidadão e ao IGPM (produto 13m).
    Ex.: aniversário abr/2026 → Index(2026-04) / Index(2025-03) − 1
    Retorna (percentual_float, 12) ou (None, 0) se dados ausentes.
    """
    key_fim = f"{data_aniversario.year}-{data_aniversario.month:02d}"
    m_ini = data_aniversario.month - 1
    a_ini = data_aniversario.year - 1
    if m_ini == 0:
        m_ini = 12
        a_ini -= 1
    key_ini = f"{a_ini}-{m_ini:02d}"
    val_fim = hist_indice.get(key_fim)
    val_ini = hist_indice.get(key_ini)
    if not val_ini or not val_fim:
        return None, 0
    return round((val_fim / val_ini - 1) * 100, 4), 12


def calcular_acumulado_12m(historico, data_aniversario):
    """Calcula o índice acumulado de 13 meses (aniversário anterior até aniversário atual, inclusive).
    Ex.: aniversário abril/2026 → janela abr/2025 a abr/2026 (13 variações mensais).
    Equivale ao cálculo do BACEN Cidadão.
    Retorna (percentual_float, meses_encontrados).
    """
    ano_ref = data_aniversario.year
    mes_ref = data_aniversario.month
    # 13 meses: de (ano_ref-1, mes_ref) até (ano_ref, mes_ref) — ambos inclusive
    keys = []
    for i in range(12, -1, -1):
        m = mes_ref - i
        a = ano_ref
        while m <= 0:
            m += 12
            a -= 1
        keys.append(f"{a}-{m:02d}")
    acc = 1.0
    found = 0
    for k in keys:
        if k in historico:
            acc *= (1 + historico[k] / 100)
            found += 1
    if found == 0:
        return None, 0
    return round((acc - 1) * 100, 4), found


def ler_excel_reajustes(path):
    """Lê a planilha de contratos e retorna lista de dicts prontos para exibição.
    Inclui 'num_linha' (1-based) para permitir escrita posterior de volta na planilha.

    Status baseado no mês do aniversário:
      ESTE_MES  → mês atual == mês SEGUINTE ao aniversário (reajustar agora)
      FUTURO    → aniversário ainda não chegou este ano
      OK        → fora da época de reajuste
    """
    wb = load_workbook(str(path), data_only=True)
    ws = wb.active
    contratos = []
    hoje = date.today()

    for num_linha, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        loc  = row[4] if len(row) > 4 else None   # col E – Locatário
        alug = row[5] if len(row) > 5 else None   # col F – Aluguel
        if not loc or not isinstance(alug, (int, float)):
            continue
        data_rej   = parse_data_reajuste(row[7]  if len(row) > 7  else None)  # col H
        data_ini   = parse_data_reajuste(row[8]  if len(row) > 8  else None)  # col I
        data_fim   = parse_data_reajuste(row[9]  if len(row) > 9  else None)  # col J
        indice_raw = row[10] if len(row) > 10 else None                        # col K
        indice_norm = normalizar_indice(str(indice_raw) if indice_raw else '')
        if not data_rej or not indice_norm:
            continue

        # Mês de aplicação = mês SEGUINTE ao aniversário
        # Ex.: aniversário setembro (9) → aplicar em outubro (10)
        #       aniversário dezembro (12) → aplicar em janeiro (1)
        mes_aplicacao = data_rej.month % 12 + 1

        if mes_aplicacao == hoje.month:
            status = 'ESTE_MES'
        elif mes_aplicacao > hoje.month:
            status = 'FUTURO'
        else:
            status = 'OK'

        # Aniversário para cálculo BACEN: sempre ano atual.
        # NÃO avança para ano seguinte — o BACEN calcula os 12 meses que PRECEDEM
        # o mês do aniversário; avançar para 2027 jogaria a janela para meses futuros
        # sem dados, resultando em "—".
        try:
            aniv = data_rej.replace(year=hoje.year)
        except ValueError:
            aniv = data_rej.replace(year=hoje.year, day=28)

        # Vigência mínima: o contrato precisa ter pelo menos 1 ano completo
        # (data_inicio até o aniversário atual >= 365 dias) para ser reajustável.
        # Se data_inicio ausente, assume-se vigência ok.
        if data_ini:
            vigencia_ok = (aniv - data_ini).days >= 365
        else:
            vigencia_ok = True

        contratos.append({
            'num_linha':         num_linha,
            'locatario':         str(loc).strip(),
            'proprietario':      str(row[3]).strip() if len(row) > 3 and row[3] else '',
            'endereco':          str(row[2]).strip() if len(row) > 2 and row[2] else '',
            'aluguel':           float(alug),
            'indice':            str(indice_raw).strip() if indice_raw else '',
            'indice_norm':       indice_norm,
            'data_reajuste':     data_rej.strftime('%d/%m'),    # só DD/MM
            'data_reajuste_iso': aniv.isoformat(),              # próxima ocorrência (usado pelo BACEN)
            'data_inicio':       data_ini.strftime('%d/%m/%Y') if data_ini else '',
            'data_fim':          data_fim.strftime('%d/%m/%Y') if data_fim else '',
            'status':            status,
            'mes_reajuste':      data_rej.month,
            'vigencia_ok':       vigencia_ok,
        })

    # Ordenar: este mês primeiro, futuros em ordem de mês, já feitos por último
    ordem = {'ESTE_MES': 0, 'FUTURO': 1, 'OK': 2}
    contratos.sort(key=lambda c: (ordem[c['status']], c['mes_reajuste']))
    return contratos


def _avancar_aniversario(d):
    """Avança a data de aniversário exatamente 1 ano."""
    try:
        return d.replace(year=d.year + 1)
    except ValueError:
        # 29/fev em ano não-bissexto → vai para 28/fev
        return d.replace(year=d.year + 1, day=28)


def aplicar_reajustes_excel(path, contratos_aplicar):
    """Escreve novos aluguéis na planilha e salva aluguel antigo no histórico DIMOB.

    contratos_aplicar: lista de dicts com ao menos:
        num_linha, novo_aluguel, locatario, mes_reajuste
    Retorna (n_atualizados, erros[])
    """
    from openpyxl import load_workbook as _lw
    wb = _lw(str(path))
    ws = wb.active
    erros = []
    n = 0

    historico_dimob = ler_dimob_historico()
    ano_atual = str(date.today().year)
    if ano_atual not in historico_dimob:
        historico_dimob[ano_atual] = {}

    for c in contratos_aplicar:
        try:
            nl        = int(c['num_linha'])
            novo_alug = float(c['novo_aluguel'])

            # Lê aluguel antigo ANTES de sobrescrever — guarda no histórico DIMOB
            aluguel_antigo = safe_float(ws.cell(row=nl, column=6).value)
            locatario      = c.get('locatario', '')
            mes_rej        = c.get('mes_reajuste')
            mes_aplicacao  = (int(mes_rej) % 12 + 1) if mes_rej else None
            chave_dimob    = norm(locatario) if locatario else f"linha_{nl}"

            historico_dimob[ano_atual][chave_dimob] = {
                'aluguel_antigo':  aluguel_antigo,
                'aluguel_novo':    novo_alug,
                'locatario':       locatario,
                'mes_aplicacao':   mes_aplicacao,
                'num_linha':       nl,
            }

            ws.cell(row=nl, column=6).value = novo_alug   # col F – único campo alterado
            n += 1
        except Exception as e:
            erros.append(f"Linha {c.get('num_linha','?')} ({c.get('locatario','')}): {e}")

    try:
        salvar_dimob_historico(historico_dimob)
    except Exception:
        pass

    # Escrita atômica
    tmp = Path(str(path) + '.tmp')
    wb.save(str(tmp))
    tmp.replace(Path(path))
    return n, erros


def similaridade(palavras_a, palavras_b):
    if not palavras_a or not palavras_b:
        return 0.0
    sa, sb = set(palavras_a), set(palavras_b)
    overlap = len(sa & sb)
    return overlap / max(len(sa), len(sb))

def is_edificio_excluido(condo):
    """Retorna True se o boleto pertence a um edificio onde só se cobra aluguel (sem repasse cond)."""
    campos = [condo.get("edificio", ""), condo.get("arquivo", "")]
    texto = " ".join(str(c) for c in campos)
    texto = unicodedata.normalize("NFKD", texto.upper())
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    for excl in EDIFICIOS_EXCLUIDOS:
        if excl.upper() in texto:
            return True
    return False


def incrementar_parcela(parcela_str):
    """'03/07' → '04/07'. Retorna None se foi a última parcela."""
    if not parcela_str:
        return None
    m = re.match(r'(\d+)/(\d+)', str(parcela_str).strip())
    if not m:
        return None
    atual, total = int(m.group(1)), int(m.group(2))
    if atual >= total:
        return None  # Última parcela — não replicar
    return f"{atual+1:02d}/{total:02d}"


# ─── Config / Histórico ────────────────────────────────────────────────────────

def salvar_config(d):
    # Escrita atômica: grava em .tmp e renomeia, evitando corrupção parcial.
    tmp = CONFIG_FILE.with_suffix(".tmp")
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(d, f, ensure_ascii=False, indent=2)
    tmp.replace(CONFIG_FILE)

def ler_config():
    cfg = {}
    if CONFIG_FILE.exists():
        try:
            with open(CONFIG_FILE, encoding="utf-8") as f:
                cfg = json.load(f)
        except Exception:
            pass
    # Variável de ambiente tem prioridade (usada na nuvem)
    env_key = os.environ.get("ANTHROPIC_API_KEY", "")
    if env_key:
        cfg["api_key"] = env_key
    return cfg

def salvar_historico(d):
    # Escrita atômica: grava em arquivo temporário e renomeia,
    # evitando corrupção caso o processo seja interrompido no meio.
    tmp = HISTORICO_FILE.with_suffix(".tmp")
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(d, f, ensure_ascii=False, indent=2)
    tmp.replace(HISTORICO_FILE)

def ler_historico():
    if HISTORICO_FILE.exists():
        try:
            with open(HISTORICO_FILE, encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            # Arquivo corrompido — começa do zero sem travar o sistema
            return {}
    return {}


# ─── DIMOB helpers ────────────────────────────────────────────────────────────

def ler_dimob_historico():
    if DIMOB_HISTORICO_FILE.exists():
        try:
            with open(DIMOB_HISTORICO_FILE, encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            return {}
    return {}

def salvar_dimob_historico(dados):
    with open(DIMOB_HISTORICO_FILE, 'w', encoding='utf-8') as f:
        json.dump(dados, f, ensure_ascii=False, indent=2)

# ─── Emails dos Locatários ────────────────────────────────────────────────────

def ler_locatarios_emails():
    if LOCATARIOS_EMAILS_FILE.exists():
        try:
            with open(LOCATARIOS_EMAILS_FILE, encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            return {}
    return {}

def salvar_locatarios_emails(d):
    tmp = LOCATARIOS_EMAILS_FILE.with_suffix('.tmp')
    with open(tmp, 'w', encoding='utf-8') as f:
        json.dump(d, f, ensure_ascii=False, indent=2)
    tmp.replace(LOCATARIOS_EMAILS_FILE)

def ler_excel_dimob(path):
    """Lê contratos para DIMOB incluindo CPF/CNPJ (colunas L=11, M=12) se existirem."""
    wb = load_workbook(str(path), data_only=True)
    ws = wb.active
    contratos = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or len(row) < 5 or not row[4]:
            continue
        locatario    = str(row[4]).strip() if row[4] else ''
        proprietario = str(row[3]).strip() if len(row) > 3 and row[3] else ''
        endereco     = str(row[2]).strip() if len(row) > 2 and row[2] else ''
        aluguel      = safe_float(row[5]) if len(row) > 5 and row[5] else 0.0
        perc_imob    = safe_float(row[6]) if len(row) > 6 and row[6] else 0.0
        data_rej     = parse_data_reajuste(row[7]) if len(row) > 7 and row[7] else None
        data_ini     = parse_data_reajuste(row[8]) if len(row) > 8 and row[8] else None
        data_fim     = parse_data_reajuste(row[9]) if len(row) > 9 and row[9] else None
        def _cpf(v):
            if not v:
                return ''
            s = str(v).strip()
            # Remove ".0" quando openpyxl lê CPF numérico como float
            if s.endswith('.0') and s[:-2].isdigit():
                s = s[:-2]
            return s
        cpf_prop = _cpf(row[11] if len(row) > 11 else None)
        cpf_loc  = _cpf(row[12] if len(row) > 12 else None)

        if not locatario or aluguel == 0:
            continue

        mes_rej       = data_rej.month if data_rej else None
        mes_aplicacao = (mes_rej % 12 + 1) if mes_rej else None

        contratos.append({
            'num_linha':       row_idx,
            'locatario':       locatario,
            'proprietario':    proprietario,
            'endereco':        endereco,
            'aluguel':         aluguel,
            'percentual_imob': perc_imob,
            'mes_reajuste':    mes_rej,
            'mes_aplicacao':   mes_aplicacao,
            'data_inicio':     data_ini,
            'data_fim':        data_fim,
            'cpf_proprietario': cpf_prop,
            'cpf_locatario':   cpf_loc,
        })
    return contratos

def calcular_meses_dimob(contrato, ano, historico_dimob):
    """Calcula aluguel de cada mês do ano de referência para DIMOB.
    Todos os 12 meses recebem um valor. Quando há reajuste registrado,
    meses anteriores ao mes_aplicacao usam aluguel_antigo e os demais usam aluguel_atual.
    """
    aluguel_atual = contrato['aluguel']
    mes_aplicacao = contrato.get('mes_aplicacao')

    chave    = norm(contrato['locatario'])
    hist_ano = historico_dimob.get(str(ano), {})
    hist_c   = hist_ano.get(chave, {})
    aluguel_antigo = hist_c.get('aluguel_antigo')
    tem_historico  = aluguel_antigo is not None and mes_aplicacao is not None

    multa_juros_hist = hist_c.get("multa_juros", {})

    meses = []
    for mes in range(1, 13):
        if tem_historico and mes < mes_aplicacao:
            base = aluguel_antigo
        else:
            base = aluguel_atual
        # Soma multa+juros do mês se houver
        mj = multa_juros_hist.get(str(mes), {})
        extra = float(mj.get("multa", 0) or 0) + float(mj.get("juros", 0) or 0)
        meses.append(round(base + extra, 2))

    return meses, tem_historico


# ─── Excel ────────────────────────────────────────────────────────────────────

def ler_excel(path):
    wb = load_workbook(path)
    ws = wb.active

    # 1ª passagem: coleta linhas válidas e conta repetições de nome
    linhas = []
    contagem_nomes = {}
    for row in ws.iter_rows(values_only=True):
        if not row[4]:
            continue
        locatario = str(row[4]).strip()
        aluguel = row[5]
        if not isinstance(aluguel, (int, float)):
            continue
        linhas.append(row)
        chave_base = norm(locatario)
        contagem_nomes[chave_base] = contagem_nomes.get(chave_base, 0) + 1

    # 2ª passagem: monta o dicionário de contratos
    # Locatários com nome duplicado recebem chave = nome + unidade,
    # garantindo que dois imóveis do mesmo inquilino apareçam separados.
    contratos = {}
    proprietarios = {}
    for row in linhas:
        locatario = str(row[4]).strip()
        aluguel    = row[5]
        proprietario = str(row[3]).strip() if row[3] else ""
        endereco     = str(row[2]).strip() if row[2] else ""
        tipo         = str(row[1]).strip() if row[1] else ""

        chave_base = norm(locatario)
        unidade    = extrair_unidade(endereco)

        if contagem_nomes[chave_base] > 1:
            # Mesmo nome em duas linhas → diferencia pela unidade do endereço
            chave = chave_base + (unidade if unidade else norm(endereco)[:8])
        else:
            chave = chave_base

        # Garante unicidade mesmo se unidade também coincidir
        sufixo = 2
        chave_final = chave
        while chave_final in contratos:
            chave_final = chave + str(sufixo)
            sufixo += 1

        # Percentual da imobiliária (coluna G — 7ª coluna, índice 6) — opcional
        try:
            perc_raw = row[6] if len(row) > 6 else None
            percentual_imob = float(perc_raw) if perc_raw is not None and str(perc_raw).strip() != "" else 0.0
        except (TypeError, ValueError):
            percentual_imob = 0.0

        email_prop = str(row[13]).strip() if len(row) > 13 and row[13] else ""

        contratos[chave_final] = {
            "locatario": locatario,
            "proprietario": proprietario,
            "endereco": endereco,
            "tipo": tipo,
            "aluguel": float(aluguel),
            "percentual_imob": percentual_imob,
            "email_proprietario": email_prop,
            "_palavras_loc": norm_palavras(locatario),
            "_palavras_prop": norm_palavras(proprietario),
            "_prop_norm": norm(proprietario),
            "_rua": extrair_rua(endereco),
            "_unidade": unidade,
        }
        chave_prop = norm(proprietario)
        if chave_prop:
            proprietarios[chave_prop] = chave_final
        for palavra in norm_palavras(proprietario):
            if len(palavra) > 4:
                proprietarios.setdefault("__w__" + palavra, chave_final)
    return contratos, proprietarios


# ─── PDF helpers ───────────────────────────────────────────────────────────────

def pdf_tem_texto(path):
    try:
        with pdfplumber.open(str(path)) as pdf:
            for page in pdf.pages:
                t = page.extract_text()
                if t and len(t.strip()) > 80:
                    return True
    except Exception:
        pass
    return False

def pdf_para_b64(path, resolucao=150):
    imgs = []
    try:
        with pdfplumber.open(str(path)) as pdf:
            for page in pdf.pages:
                img = page.to_image(resolution=resolucao)
                buf = io.BytesIO()
                img.save(buf, format="PNG")
                buf.seek(0)
                imgs.append(base64.b64encode(buf.read()).decode("utf-8"))
    except Exception:
        pass
    return imgs

def pdf_texto_completo(path):
    t = ""
    try:
        with pdfplumber.open(str(path)) as pdf:
            for page in pdf.pages:
                txt = page.extract_text()
                if txt:
                    t += txt + "\n"
    except Exception:
        pass
    return t.strip()

def montar_content(pdf_path, prompt):
    if pdf_tem_texto(pdf_path):
        txt = pdf_texto_completo(pdf_path)
        return [{"type": "text", "text": f"Texto extraido do PDF:\n\n{txt}\n\n{prompt}"}]
    imgs = pdf_para_b64(pdf_path)
    if not imgs:
        return None
    content = []
    for img in imgs[:3]:
        content.append({"type": "image",
                         "source": {"type": "base64", "media_type": "image/png", "data": img}})
    content.append({"type": "text", "text": prompt})
    return content


def chamar_claude(api_key, content, max_tokens=1500, tentativas=3):
    client = anthropic.Anthropic(api_key=api_key)
    ultimo_erro = "Falhou apos todas as tentativas"
    for t in range(tentativas):
        try:
            resp = client.messages.create(
                model="claude-haiku-4-5-20251001",
                max_tokens=max_tokens,
                messages=[{"role": "user", "content": content}]
            )
            texto = resp.content[0].text.strip()
            m = re.search(r"\{[\s\S]*\}", texto)
            if m:
                return json.loads(m.group()), None
            # Resposta sem JSON — pode ser falha transitória, tenta de novo
            ultimo_erro = f"Sem JSON na resposta: {texto[:120]}"
            if t < tentativas - 1:
                time.sleep(2)
                continue
            return None, ultimo_erro
        except json.JSONDecodeError as e:
            ultimo_erro = f"JSON invalido: {e}"
            if t < tentativas - 1:
                time.sleep(1)
                continue
            return None, ultimo_erro
        except anthropic.RateLimitError as e:
            # Limite de requisições — aguarda mais e tenta novamente
            ultimo_erro = f"Rate limit: {e}"
            time.sleep(5 * (t + 1))
            continue
        except anthropic.APIStatusError as e:
            # Erro HTTP da API (503, 529 overloaded, etc.) — transitório
            ultimo_erro = f"API status {e.status_code}: {e.message}"
            if t < tentativas - 1:
                time.sleep(4 * (t + 1))
                continue
            return None, ultimo_erro
        except anthropic.BadRequestError as e:
            # Conteúdo inválido — erro permanente, não adianta tentar de novo
            return None, f"Conteudo rejeitado pela API: {e}"
        except Exception as e:
            ultimo_erro = str(e)
            if t < tentativas - 1:
                time.sleep(2)
                continue
            return None, ultimo_erro
    return None, ultimo_erro


# ─── Extração condomínio ───────────────────────────────────────────────────────

PROMPT_CONDO = """Analise este documento de condominio brasileiro.

PRIMEIRO, identifique o tipo:
- "boleto_individual": boleto de cobranca de UMA unidade especifica
- "demonstrativo_com_recibo": documento que contem balancete do predio E cobranca individual da unidade
- "demonstrativo_geral": balancete/demonstrativo do PREDIO INTEIRO sem dados individuais por unidade

Retorne APENAS JSON valido:
{
  "tipo_documento": "boleto_individual" ou "demonstrativo_com_recibo" ou "demonstrativo_geral",
  "edificio": "nome do condominio/edificio",
  "endereco": "rua e numero do imovel (do predio, nao do pagador)",
  "unidade": "numero da unidade/apartamento sem prefixo (ex: 56A, 37, 103B)",
  "pagador": "nome completo do proprietario/pagador",
  "ac": "nome do locatario no campo A/C ou Ao cuidado de, ou null",
  "mes_referencia": "mes/ano ex: abril/2026",
  "itens": {
    "cota": null,
    "agua": null,
    "gas": null,
    "energia": null,
    "tx_leitura": null,
    "vaga_moto": null,
    "tag": null,
    "melhorias": null,
    "fundo_reserva": null,
    "correio": null,
    "outros": null
  },
  "total_boleto": 0
}

ATENCAO:
- Se o documento tiver demonstrativo do predio inteiro E dados individuais da unidade: tipo = demonstrativo_com_recibo
- Se tiver so demonstrativo do predio sem dados individuais: tipo = demonstrativo_geral
- Se tiver so boleto da unidade: tipo = boleto_individual

REGRAS CRITICAS DE CLASSIFICACAO — LEIA COM MAXIMA ATENCAO:

AGUA (campo "agua"):
- Qualquer linha com as palavras: "Agua", "Agua Fria", "Consumo Agua", "Cons. Agua", "Cons Agua", "Consumo de Agua",
  "Hidrometro", "Leitura Hidrometro", "Taxa Agua", "Agua Individual", "Medicao Agua"
- E um valor VARIAVEL que muda todo mes conforme o consumo medido
- Valor tipico: R$10 a R$300
- SE HOUVER QUALQUER LINHA COM ESSAS PALAVRAS, coloque o valor em "agua". NUNCA em "fundo_reserva".

FUNDO DE RESERVA (campo "fundo_reserva"):
- Somente linhas que contenham EXATAMENTE: "Fundo de Reserva", "Fundo Reserva", "F. Reserva", "FR", "Reserva"
- E um percentual FIXO da cota (geralmente 5% a 20% da cota mensal)
- SE NAO TIVER A PALAVRA "Reserva" ou "Fundo" na linha, NAO e fundo_reserva.

REGRA DE OURO: na duvida entre agua e fundo_reserva, use "agua" se o valor for variavel/consumo medido.

OUTRAS CLASSIFICACOES:
- "tx_leitura" = taxa fixa de leitura de hidrometro (separado do consumo).
- "cota" = quota condominial ordinaria mensal (valor fixo basico).
- "gas" = consumo de gas medido (valor variavel mensal).
- "energia" = energia eletrica de areas comuns.
- "melhorias" = obras, reformas, benfeitorias.
- "correio" = correspondencia/malote.
- "vaga_moto" = taxa de vaga de garagem ou moto.
- "tag" = tag de acesso, controle de acesso.
- "outros" = qualquer encargo nao classificado acima.

- Valores com ponto decimal (1234.56). null para ausentes.
- Para demonstrativo_geral sem dados individuais: deixe todos os itens null."""

def extrair_condo(api_key, pdf_path):
    content = montar_content(pdf_path, PROMPT_CONDO)
    if not content:
        return None, "Nao converteu PDF"
    dados, erro = chamar_claude(api_key, content)
    if dados:
        dados["arquivo"] = Path(pdf_path).name
        return dados, None
    return None, erro or "Falhou"


# ─── Extração boleto anterior ─────────────────────────────────────────────────

PROMPT_BOLETO_ANT = """Analise este boleto de locacao/aluguel do mes anterior. Extraia TODOS os valores cobrados.

Retorne APENAS JSON valido:
{
  "locatario": "nome completo do locatario",
  "endereco": "endereco do imovel alugado",
  "aluguel": null,
  "cond_cota": null,
  "cond_vaga": null,
  "agua": null,
  "gas": null,
  "energia": null,
  "tx_leitura": null,
  "iptu": null,
  "iptu_parcela": "formato parcela/total ex: 03/07 (NUNCA mes/ano). null se ausente.",
  "iptu_vaga": null,
  "iptu_vaga_parcela": "formato parcela/total ex: 03/06 (NUNCA mes/ano). null se ausente.",
  "seg_fianca": null,
  "seg_incendio": null,
  "outros_itens": [{"descricao": "nome", "valor": 0.0}]
}

Instrucoes:
- "cond_cota" = taxa de condominio (valor fixo mensal)
- "cond_vaga" = taxa de vaga de garagem/condominio vaga
- "agua", "gas", "energia" = consumos variaveis
- "iptu" e "iptu_vaga" = IPTU do apartamento e da vaga (podem ser cobrados separados)
- iptu_parcela e iptu_vaga_parcela: SOMENTE formato "parcela/total" ex: "03/07". NUNCA colocar mes ou ano.
- Valores com ponto decimal. null para ausentes."""

def extrair_boleto_anterior(api_key, pdf_path):
    content = montar_content(pdf_path, PROMPT_BOLETO_ANT)
    if not content:
        return None, "Nao converteu"
    dados, erro = chamar_claude(api_key, content)
    if dados:
        dados["arquivo"] = Path(pdf_path).name
        return dados, None
    return None, erro or "Falhou"


# ─── Matching ─────────────────────────────────────────────────────────────────

def match_condo_locatario(condo, contratos, proprietarios):
    candidatos = {}

    ac = condo.get("ac") or ""
    pagador = condo.get("pagador") or ""
    condo_rua = extrair_rua(condo.get("endereco", ""))
    condo_unidade = re.sub(r"[^0-9A-Z]", "", str(condo.get("unidade", "")).upper())

    # Extrai nome e unidade do nome do arquivo (colocados manualmente pelo usuário)
    nome_arq, unidade_arq = extrair_nome_arquivo(condo.get("arquivo", ""))
    # Usa unidade do filename como fallback quando a do PDF não veio ou parece errada
    if unidade_arq and not condo_unidade:
        condo_unidade = unidade_arq

    # 1. A/C exato
    if ac:
        chave = norm(ac)
        if chave in contratos:
            return chave, "ac_exato", 100

    # 2. A/C por palavras — usa score_nome_arquivo (precision) para aceitar nomes parciais ex: "Denis"
    if ac:
        palavras_ac = norm_palavras(ac)
        for chave, ct in contratos.items():
            s = score_nome_arquivo(palavras_ac, ct["_palavras_loc"])
            if s >= 0.6:
                score = s * 98
                if score > candidatos.get(chave, (0, ""))[0]:
                    candidatos[chave] = (score, "ac_palavras")

    # 3. Nome do arquivo — compara contra locatário E proprietário
    #    (muitos arquivos são nomeados pelo proprietário, não pelo locatário)
    if nome_arq:
        # Extrai palavras do nome do arquivo aceitando abreviações de 2+ letras (ex: HP)
        _n = unicodedata.normalize("NFKD", nome_arq.upper())
        _n = "".join(c for c in _n if not unicodedata.combining(c))
        palavras_arq = [w for w in re.sub(r"[^A-Z0-9 ]", " ", _n).split() if len(w) >= 2]
        for chave, ct in contratos.items():
            # Bonus de rua: quando a rua do boleto bate com a do contrato
            # resolve empates (ex: dois locatários com mesmo proprietário "João")
            bonus_rua = 6 if (condo_rua and ct["_rua"] == condo_rua) else 0

            # Tenta contra o locatário primeiro (score mais alto)
            s = score_nome_arquivo(palavras_arq, ct["_palavras_loc"])
            if s >= 0.5:
                score = s * 93 + bonus_rua
                if score > candidatos.get(chave, (0, ""))[0]:
                    candidatos[chave] = (score, "nome_arquivo_loc")
            # Tenta contra o proprietário (score menor — nome no arquivo é do locatário)
            s2 = score_nome_arquivo(palavras_arq, ct["_palavras_prop"])
            if s2 >= 0.6:
                bonus_un = 10 if (unidade_arq and ct.get("_unidade") == unidade_arq) else 0
                score = s2 * 87 + bonus_rua + bonus_un
                if score > candidatos.get(chave, (0, ""))[0]:
                    candidatos[chave] = (score, "nome_arquivo_prop")
            # Abreviação curta (2 chars como HP): compara diretamente com nome normalizado
            _nome_norm = norm(nome_arq)
            if len(_nome_norm) >= 2 and _nome_norm == ct["_prop_norm"]:
                score = 85 + bonus_rua
                if score > candidatos.get(chave, (0, ""))[0]:
                    candidatos[chave] = (score, "nome_arquivo_prop_abbrev")

    # 4. Proprietário exato — itera TODOS os contratos com aquele proprietário
    #    (o dict proprietarios guarda só um; iterando todos resolvemos casos
    #     onde dois locatários têm o mesmo proprietário, ex: dois imóveis do João)
    if pagador:
        chave_prop = norm(pagador)
        if chave_prop in proprietarios:
            for chave, ct in contratos.items():
                if ct["_prop_norm"] == chave_prop:
                    bonus = 6 if (condo_rua and ct.get("_rua") == condo_rua) else 0
                    bonus += 8 if (condo_unidade and ct.get("_unidade") == condo_unidade) else 0
                    score = 90 + bonus
                    if score > candidatos.get(chave, (0, ""))[0]:
                        candidatos[chave] = (score, "prop_exato")

    # 5. Proprietário por palavras — também com bonus de rua
    if pagador:
        palavras_pag = norm_palavras(pagador)
        for chave, ct in contratos.items():
            s = similaridade(palavras_pag, ct["_palavras_prop"])
            if s >= 0.5:
                bonus = 6 if (condo_rua and ct.get("_rua") == condo_rua) else 0
                score = s * 85 + bonus
                if score > candidatos.get(chave, (0, ""))[0]:
                    candidatos[chave] = (score, "prop_palavras")

    # 6. Endereço (rua) + unidade — principal para VM/CV (Mandaqui/Itauna)
    if condo_rua and condo_unidade:
        for chave, ct in contratos.items():
            if ct["_rua"] == condo_rua and ct["_unidade"] == condo_unidade:
                score = 92
                if score > candidatos.get(chave, (0, ""))[0]:
                    candidatos[chave] = (score, "rua+unidade")

    # 7. Rua + unidade do filename (quando a unidade do PDF estava errada)
    if condo_rua and unidade_arq and unidade_arq != condo_unidade:
        for chave, ct in contratos.items():
            if ct["_rua"] == condo_rua and ct["_unidade"] == unidade_arq:
                score = 88
                if score > candidatos.get(chave, (0, ""))[0]:
                    candidatos[chave] = (score, "rua+unidade_arquivo")

    # 8. Só unidade do filename quando rua não veio no PDF
    if unidade_arq and not condo_rua:
        matches_unidade = [(k, ct) for k, ct in contratos.items()
                           if ct["_unidade"] == unidade_arq]
        if len(matches_unidade) == 1:
            chave = matches_unidade[0][0]
            score = 65
            if score > candidatos.get(chave, (0, ""))[0]:
                candidatos[chave] = (score, "unidade_arquivo_unica")

    # 9. Só unidade do PDF (quando rua não veio)
    if condo_unidade and not condo_rua:
        matches_unidade = [(k, ct) for k, ct in contratos.items()
                           if ct["_unidade"] == condo_unidade]
        if len(matches_unidade) == 1:
            chave = matches_unidade[0][0]
            score = 65
            if score > candidatos.get(chave, (0, ""))[0]:
                candidatos[chave] = (score, "unidade_unica")

    # 10. Palavra individual do proprietário (score=35 < mínimo 50 = nunca vence sozinha)
    if pagador:
        for palavra in norm_palavras(pagador):
            if len(palavra) > 5:
                key = "__w__" + palavra
                if key in proprietarios:
                    chave_loc = proprietarios[key]
                    score = 35
                    if score > candidatos.get(chave_loc, (0, ""))[0]:
                        candidatos[chave_loc] = (score, "prop_parcial")

    if not candidatos:
        return None, None, 0
    melhor = max(candidatos, key=lambda k: candidatos[k][0])
    score, metodo = candidatos[melhor]
    if score < 50:
        return None, None, 0
    return melhor, metodo, int(score)


def match_boleto_locatario(boleto, contratos):
    nome = boleto.get("locatario", "")
    if not nome:
        return None
    chave = norm(nome)
    if chave in contratos:
        return chave
    palavras = norm_palavras(nome)
    melhor, melhor_s = None, 0
    for chave, ct in contratos.items():
        s = similaridade(palavras, ct["_palavras_loc"])
        if s > melhor_s:
            melhor_s = s
            melhor = chave
    return melhor if melhor_s >= 0.6 else None


# ─── Flask ────────────────────────────────────────────────────────────────────

@app.route("/login", methods=["GET", "POST"])
def login():
    erro = ""
    if request.method == "POST":
        usuario = request.form.get("usuario", "").strip()
        senha   = request.form.get("senha", "").strip()
        if usuario in SENHAS and senha == SENHAS[usuario]:
            session["logado"] = True
            session["role"]   = USUARIOS[usuario]
            session["usuario"] = usuario
            return redirect(url_for("home"))
        erro = "Usuário ou senha incorretos."
    return render_template("login.html", erro=erro)

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))

@app.route("/")
@login_required
def home():
    return render_template("home.html")

@app.route("/boletos")
@login_required
def index():
    config = ler_config()
    return render_template("index.html", tem_api_key=bool(config.get("api_key")), is_admin=is_admin())

@app.route("/config", methods=["POST"])
@admin_required
def config_save():
    api_key = (request.json or {}).get("api_key", "").strip()
    if not api_key:
        return jsonify({"ok": False, "erro": "Chave vazia"})
    try:
        client = anthropic.Anthropic(api_key=api_key)
        client.messages.create(model="claude-haiku-4-5-20251001", max_tokens=10,
                                messages=[{"role": "user", "content": "ok"}])
    except Exception as e:
        return jsonify({"ok": False, "erro": str(e)[:150]})
    salvar_config({"api_key": api_key})
    return jsonify({"ok": True})

@app.route("/processar", methods=["POST"])
@login_required
def processar():
    config = ler_config()
    api_key = config.get("api_key")
    if not api_key:
        return jsonify({"ok": False, "erro": "Configure a chave API primeiro"})

    excel_file = request.files.get("excel")
    condo_files = request.files.getlist("condominios")
    boleto_files = request.files.getlist("boletos_anteriores")

    if not excel_file:
        return jsonify({"ok": False, "erro": "Envie o arquivo Excel"})

    if UPLOAD_DIR.exists():
        for f in UPLOAD_DIR.iterdir():
            try:
                if f.is_file() and f.suffix.lower() != '.xlsx':
                    f.unlink()
            except Exception:
                pass
    UPLOAD_DIR.mkdir(exist_ok=True)

    excel_path = UPLOAD_DIR / "contratos.xlsx"
    excel_file.save(str(excel_path))

    condo_paths = []
    for f in condo_files:
        if f.filename.lower().endswith(".pdf"):
            safe_name = Path(f.filename).name  # sanitiza path traversal
            dest = UPLOAD_DIR / ("cond_" + safe_name)
            f.save(str(dest))
            condo_paths.append(dest)

    boleto_paths = []
    for f in boleto_files:
        if f.filename.lower().endswith(".pdf"):
            safe_name = Path(f.filename).name  # sanitiza path traversal
            dest = UPLOAD_DIR / ("bol_" + safe_name)
            f.save(str(dest))
            boleto_paths.append(dest)

    def gerar():
        total = len(condo_paths) + len(boleto_paths)
        yield f"data: {json.dumps({'tipo': 'inicio', 'total': total})}\n\n"

        try:
            contratos, proprietarios = ler_excel(excel_path)
            yield f"data: {json.dumps({'tipo': 'log', 'msg': f'Excel: {len(contratos)} locatarios'})}\n\n"
            # Salva nomes dos locatários para o módulo de envio de boletos
            try:
                nomes = {norm(ct["locatario"]): ct["locatario"] for ct in contratos.values()}
                with open(DATA_DIR / "locatarios_nomes.json", "w", encoding="utf-8") as _f:
                    json.dump(nomes, _f, ensure_ascii=False, indent=2)
            except Exception:
                pass
        except Exception as e:
            yield f"data: {json.dumps({'tipo': 'erro', 'msg': f'Erro Excel: {e}'})}\n\n"
            return

        # Condomínios
        condos_lidos = []
        for i, pdf_path in enumerate(condo_paths):
            nome = pdf_path.name
            yield f"data: {json.dumps({'tipo': 'progresso', 'atual': i+1, 'total': total, 'arquivo': nome})}\n\n"
            dados, erro = extrair_condo(api_key, pdf_path)
            if erro:
                yield f"data: {json.dumps({'tipo': 'log', 'msg': f'AVISO {nome}: {erro}'})}\n\n"
                condos_lidos.append({"arquivo": nome, "erro": erro, "itens": {}, "locatario_chave": None, "tipo_documento": "erro"})
                continue

            tipo_doc = dados.get("tipo_documento", "boleto_individual")
            chave, metodo, conf = match_condo_locatario(dados, contratos, proprietarios)
            dados["locatario_chave"] = chave
            dados["match_metodo"] = metodo
            dados["match_confianca"] = conf

            if chave:
                loc = contratos[chave]["locatario"]
                flag = " [DEMONSTRATIVO]" if tipo_doc == "demonstrativo_geral" else (" [DEMONSTRATIVO+RECIBO]" if tipo_doc == "demonstrativo_com_recibo" else "")
                from_excluido = is_edificio_excluido(dados)
                excl_flag = " [SEM COND - EDIFICIO EXCLUIDO]" if from_excluido else ""
                yield f"data: {json.dumps({'tipo': 'log', 'msg': f'OK [{conf}%] {nome} -> {loc} ({metodo}){flag}{excl_flag}'})}\n\n"
                _itens = dados.get('itens') or {}
                _itens_str = ' | '.join(f'{k}:{v}' for k,v in _itens.items() if v) or '(todos null)'
                yield f"data: {json.dumps({'tipo': 'log', 'msg': 'ITENS ' + nome + ': ' + _itens_str})}\n\n"
            else:
                pag = dados.get("pagador", "?")
                un = dados.get("unidade", "")
                end = dados.get("endereco", "")
                yield f"data: {json.dumps({'tipo': 'log', 'msg': f'SEM MATCH {nome} | pag:{pag} | un:{un} | end:{end}'})}\n\n"
            condos_lidos.append(dados)

        # Boletos anteriores
        boletos_extras = {}
        for i, pdf_path in enumerate(boleto_paths):
            nome = pdf_path.name
            idx = len(condo_paths) + i + 1
            yield f"data: {json.dumps({'tipo': 'progresso', 'atual': idx, 'total': total, 'arquivo': nome})}\n\n"
            dados, erro = extrair_boleto_anterior(api_key, pdf_path)
            if erro:
                yield f"data: {json.dumps({'tipo': 'log', 'msg': f'AVISO boleto {nome}: {erro}'})}\n\n"
                continue
            chave = match_boleto_locatario(dados, contratos)
            if chave:
                loc = contratos[chave]["locatario"]
                boletos_extras[chave] = dados
                yield f"data: {json.dumps({'tipo': 'log', 'msg': f'OK Boleto: {nome} -> {loc}'})}\n\n"
            else:
                yield f"data: {json.dumps({'tipo': 'log', 'msg': f'SEM MATCH boleto: {nome}'})}\n\n"

        try:
            resultado = montar_resultado(contratos, condos_lidos, boletos_extras)

            historico = ler_historico()
            for row in resultado["locatarios"]:
                h = historico.get(row["chave"], {})
                if not row.get("iptu") and h.get("iptu"):
                    nova_parc = incrementar_parcela(h.get("iptu_parcela", ""))
                    if nova_parc is not None:
                        row["iptu_hist"] = h["iptu"]
                        row["iptu_parcela_hist"] = nova_parc
                if not row.get("iptu_vaga") and h.get("iptu_vaga"):
                    nova_parc_v = incrementar_parcela(h.get("iptu_vaga_parcela", ""))
                    if nova_parc_v is not None:
                        row["iptu_vaga_hist"] = h["iptu_vaga"]
                        row["iptu_vaga_parcela_hist"] = nova_parc_v
                if not row.get("seg_fianca") and h.get("seg_fianca"):
                    row["seg_fianca_hist"] = h["seg_fianca"]
                if not row.get("seg_incendio") and h.get("seg_incendio"):
                    row["seg_incendio_hist"] = h["seg_incendio"]
                # Outros fixo 1-5: replicam todo mês automaticamente
                for _n in ['', '2', '3', '4', '5']:
                    k = f"extras_fixo{_n}_val"
                    if h.get(k):
                        row[f"extras_fixo{_n}_val_hist"]  = h[k]
                        row[f"extras_fixo{_n}_desc_hist"] = h.get(f"extras_fixo{_n}_desc", "")
                # Abono: replica com incremento de parcela (igual ao IPTU)
                if h.get("abono_val"):
                    nova_parc_ab = incrementar_parcela(h.get("abono_parcela", ""))
                    if nova_parc_ab is not None:
                        row["abono_val_hist"] = h["abono_val"]
                        row["abono_parcela_hist"] = nova_parc_ab
                        row["abono_desc_hist"] = h.get("abono_desc", "")
                # Deduções manuais do repasse: replica todo mês
                for _n in range(1, 11):
                    if h.get(f"ded{_n}_val"):
                        row[f"ded{_n}_val_hist"]  = h[f"ded{_n}_val"]
                        row[f"ded{_n}_desc_hist"] = h.get(f"ded{_n}_desc", "")
                # Cota: replica do histórico somente se o boleto atual não trouxe cota
                v_cota = float(h.get("cond_cota") or 0)
                if v_cota and not row["cond_itens_rep"].get("cota"):
                    row["cond_cota_hist"] = v_cota
                    row["cond_repasse"] = round(row["cond_repasse"] + v_cota, 2)
                    row["cond_itens_rep"]["cota"] = v_cota
                # Alerta: cota condominial diferente do registrado no histórico
                cota_hist  = float(h.get("cond_cota") or 0)
                cota_atual = float((row.get("cond_itens_rep") or {}).get("cota") or 0)
                if cota_hist > 0 and cota_atual > 0 and abs(cota_atual - cota_hist) > 0.01:
                    row["cond_alerta"]     = True
                    row["cond_cota_atual"] = cota_atual
                    row["cond_cota_hist"]  = cota_hist

            yield f"data: {json.dumps({'tipo': 'resultado', 'dados': resultado})}\n\n"

        except Exception as e:
            yield f"data: {json.dumps({'tipo': 'erro', 'msg': f'Erro ao montar resultado: {e} | {traceback.format_exc()[-300:]}'})}\n\n"

    return Response(
        stream_with_context(gerar()),
        mimetype="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",
            "Connection": "keep-alive",
        }
    )


def montar_resultado(contratos, condos_lidos, boletos_extras):
    loc_condos = {}
    sem_match = []
    for c in condos_lidos:
        chave = c.get("locatario_chave")
        if chave:
            loc_condos.setdefault(chave, []).append(c)
        else:
            sem_match.append(c)

    locatarios = []
    for chave, ct in contratos.items():
        cond_list = loc_condos.get(chave, [])
        boleto_ant = boletos_extras.get(chave, {})

        # Filtro de confiança: se há múltiplos matches, descartar os de baixa confiança
        # quando existe pelo menos um de alta confiança (evita falsos positivos acumulados)
        if len(cond_list) > 1:
            max_conf = max(c.get("match_confianca", 0) for c in cond_list)
            if max_conf >= 65:
                cond_list = [c for c in cond_list if c.get("match_confianca", 0) >= max_conf - 10]

        cond_repasse = 0.0
        cond_nao_repasse = 0.0
        cond_itens_rep = {}
        cond_itens_nrep = {}
        cond_info = []

        for cond in cond_list:
            tipo_doc = cond.get("tipo_documento", "boleto_individual")
            # Cópia defensiva — evita que qualquer modificação local altere os dados originais do cond
            itens = dict(cond.get("itens") or {})
            rep = nrep = 0.0
            excluido = is_edificio_excluido(cond)

            arq_lower = cond.get("arquivo", "").lower()

            # Normaliza o nome do edifício extraído do PDF (sem acentos) para comparação
            edi_raw  = (cond.get("edificio", "") or "").lower()
            edi_norm = unicodedata.normalize("NFKD", edi_raw)
            edi_norm = "".join(c for c in edi_norm if not unicodedata.combining(c))
            end_raw  = (cond.get("endereco", "") or "").lower()
            end_norm = unicodedata.normalize("NFKD", end_raw)
            end_norm = "".join(c for c in end_norm if not unicodedata.combining(c))

            # Reconhece prédio com consumo individual por:
            # 1. "cv" ou "vm" como palavra no nome do arquivo
            #    (cobre "cv-", "cv ", "cv.", "cv_" etc.)
            # 2. Nome/endereço do edifício extraído do PDF
            #    (cobre PDFs no nome do proprietário sem cv/vm no arquivo)
            tem_consumos = (
                bool(re.search(r'\b(cv|vm)\b', arq_lower)) or
                any(k in edi_norm for k in EDIFICIOS_CONSUMO_KEYWORDS) or
                any(k in end_norm for k in EDIFICIOS_CONSUMO_KEYWORDS)
            )

            # Fora de cv_/vm_: sem consumos variáveis — só aluguel+IPTU
            if not tem_consumos:
                cond_info.append({
                    "edificio": cond.get("edificio", ""),
                    "unidade": cond.get("unidade", ""),
                    "arquivo": cond.get("arquivo", ""),
                    "tipo_documento": tipo_doc,
                    "excluido": excluido,
                    "total_boleto": safe_float(cond.get("total_boleto")),
                    "repasse": 0.0,
                    "nao_repasse": 0.0,
                    "itens": {k: safe_float(v) for k, v in itens.items() if v},
                    "match_metodo": cond.get("match_metodo", ""),
                    "match_confianca": cond.get("match_confianca", 0),
                    "rejeitado": "sem_consumos",
                })
                continue

            # ── A partir daqui: apenas cv_ e vm_ ──────────────────────────────

            # Valores grotescos (> 2000) = demonstrativo de prédio inteiro mal identificado
            tem_valor_absurdo = any(
                safe_float(itens.get(campo)) > 2000
                for campo in list(REPASSE_ITENS) + list(NAO_REPASSE)
            )
            if tem_valor_absurdo:
                cond_info.append({
                    "edificio": cond.get("edificio", ""),
                    "unidade": cond.get("unidade", ""),
                    "arquivo": cond.get("arquivo", ""),
                    "tipo_documento": tipo_doc,
                    "excluido": excluido,
                    "total_boleto": safe_float(cond.get("total_boleto")),
                    "repasse": 0.0,
                    "nao_repasse": 0.0,
                    "itens": {k: safe_float(v) for k, v in itens.items() if v},
                    "match_metodo": cond.get("match_metodo", ""),
                    "match_confianca": cond.get("match_confianca", 0),
                    "rejeitado": "valor_absurdo",
                })
                continue

            # demonstrativo_geral puro: ignora itens (valores são do prédio inteiro)
            # boleto_individual e demonstrativo_com_recibo: processa agua/gas/energia normalmente
            if tipo_doc in ("boleto_individual", "demonstrativo_com_recibo"):
                for campo in REPASSE_ITENS:
                    v = safe_float(itens.get(campo))
                    if v == 0:
                        continue
                    rep += v
                    cond_itens_rep[campo] = round(cond_itens_rep.get(campo, 0) + v, 2)

                for campo in NAO_REPASSE:
                    v = safe_float(itens.get(campo))
                    if v:
                        nrep += v
                        cond_itens_nrep[campo] = round(cond_itens_nrep.get(campo, 0) + v, 2)

            cond_repasse += rep
            cond_nao_repasse += nrep
            cond_info.append({
                "edificio": cond.get("edificio", ""),
                "unidade": cond.get("unidade", ""),
                "arquivo": cond.get("arquivo", ""),
                "tipo_documento": tipo_doc,
                "excluido": excluido,
                "total_boleto": safe_float(cond.get("total_boleto")),
                "repasse": round(rep, 2),
                "nao_repasse": round(nrep, 2),
                "itens": {k: safe_float(v) for k, v in itens.items() if v},
                "match_metodo": cond.get("match_metodo", ""),
                "match_confianca": cond.get("match_confianca", 0),
            })

        # IPTU e seguros: do boleto anterior com incremento de parcela
        iptu = 0.0
        iptu_parcela = ""
        iptu_vaga = 0.0
        iptu_vaga_parcela = ""
        seg_fianca = 0.0
        seg_incendio = 0.0
        fonte_extras = ""

        if boleto_ant:
            # IPTU do apartamento
            raw_iptu = safe_float(boleto_ant.get("iptu"))
            raw_parc = boleto_ant.get("iptu_parcela") or ""
            nova_parc = incrementar_parcela(raw_parc)
            if raw_iptu and nova_parc is not None:
                iptu = raw_iptu
                iptu_parcela = nova_parc
            elif raw_iptu and not raw_parc:
                # Sem info de parcela — mantém por segurança
                iptu = raw_iptu
                iptu_parcela = ""

            # IPTU da vaga
            raw_iptu_v = safe_float(boleto_ant.get("iptu_vaga"))
            raw_parc_v = boleto_ant.get("iptu_vaga_parcela") or ""
            nova_parc_v = incrementar_parcela(raw_parc_v)
            if raw_iptu_v and nova_parc_v is not None:
                iptu_vaga = raw_iptu_v
                iptu_vaga_parcela = nova_parc_v
            elif raw_iptu_v and not raw_parc_v:
                iptu_vaga = raw_iptu_v

            seg_fianca  = safe_float(boleto_ant.get("seg_fianca"))
            seg_incendio = safe_float(boleto_ant.get("seg_incendio"))
            if any([iptu, iptu_vaga, seg_fianca, seg_incendio]):
                fonte_extras = "boleto_anterior"

        # ── Repasse ao proprietário ───────────────────────────────────────────
        percentual_imob = ct.get("percentual_imob", 0) or 0
        taxa_imob       = round(ct["aluguel"] * percentual_imob / 100, 2)
        # cond_nao_repasse já contém fundo_reserva + correio + melhorias
        repasse_proprietario = round(ct["aluguel"] - taxa_imob - cond_nao_repasse, 2)

        locatarios.append({
            "chave": chave,
            "locatario": ct["locatario"],
            "proprietario": ct["proprietario"],
            "email_proprietario": ct.get("email_proprietario", ""),
            "endereco": ct["endereco"],
            "tipo": ct["tipo"],
            "aluguel": ct["aluguel"],
            "cond_repasse": round(cond_repasse, 2),
            "cond_nao_repasse": round(cond_nao_repasse, 2),
            "cond_itens_rep": cond_itens_rep,
            "cond_itens_nrep": cond_itens_nrep,
            "cond_info": cond_info,
            "tem_cond": bool(cond_list),
            "percentual_imob": percentual_imob,
            "taxa_imob": taxa_imob,
            "repasse_proprietario": repasse_proprietario,
            "iptu": round(iptu, 2),
            "iptu_parcela": iptu_parcela,
            "iptu_vaga": round(iptu_vaga, 2),
            "iptu_vaga_parcela": iptu_vaga_parcela,
            "seg_fianca": round(seg_fianca, 2),
            "seg_incendio": round(seg_incendio, 2),
            "fonte_extras": fonte_extras,
            "extras_desc": "",  "extras_val": 0,
            "extras2_desc": "", "extras2_val": 0,
            "extras3_desc": "", "extras3_val": 0,
            "extras_fixo_desc": "",  "extras_fixo_val": 0,
            "extras_fixo2_desc": "", "extras_fixo2_val": 0,
            "extras_fixo3_desc": "", "extras_fixo3_val": 0,
            "extras_fixo4_desc": "", "extras_fixo4_val": 0,
            "extras_fixo5_desc": "", "extras_fixo5_val": 0,
            "abono_desc": "",
            "abono_val": 0,
            "abono_parcela": "",
            "iptu_hist": 0, "iptu_parcela_hist": "",
            "iptu_vaga_hist": 0, "iptu_vaga_parcela_hist": "",
            "seg_fianca_hist": 0, "seg_incendio_hist": 0,
            "cond_cota_hist": 0,
            "extras_fixo_desc_hist": "",  "extras_fixo_val_hist": 0,
            "extras_fixo2_desc_hist": "", "extras_fixo2_val_hist": 0,
            "extras_fixo3_desc_hist": "", "extras_fixo3_val_hist": 0,
            "extras_fixo4_desc_hist": "", "extras_fixo4_val_hist": 0,
            "extras_fixo5_desc_hist": "", "extras_fixo5_val_hist": 0,
            "abono_val_hist": 0, "abono_parcela_hist": "", "abono_desc_hist": "",
            # Deduções manuais do repasse ao proprietário (10 campos)
            "ded1_desc": "", "ded1_val": 0, "ded1_desc_hist": "", "ded1_val_hist": 0,
            "ded2_desc": "", "ded2_val": 0, "ded2_desc_hist": "", "ded2_val_hist": 0,
            "ded3_desc": "", "ded3_val": 0, "ded3_desc_hist": "", "ded3_val_hist": 0,
            "ded4_desc": "", "ded4_val": 0, "ded4_desc_hist": "", "ded4_val_hist": 0,
            "ded5_desc": "", "ded5_val": 0, "ded5_desc_hist": "", "ded5_val_hist": 0,
            "ded6_desc": "", "ded6_val": 0, "ded6_desc_hist": "", "ded6_val_hist": 0,
            "ded7_desc": "", "ded7_val": 0, "ded7_desc_hist": "", "ded7_val_hist": 0,
            "ded8_desc": "", "ded8_val": 0, "ded8_desc_hist": "", "ded8_val_hist": 0,
            "ded9_desc": "", "ded9_val": 0, "ded9_desc_hist": "", "ded9_val_hist": 0,
            "ded10_desc": "", "ded10_val": 0, "ded10_desc_hist": "", "ded10_val_hist": 0,
        })

    locatarios.sort(key=lambda x: x["locatario"])
    return {"locatarios": locatarios, "sem_match": sem_match}


# ─── SMTP ─────────────────────────────────────────────────────────────────────

def ler_smtp():
    cfg = ler_config()
    return {
        "host":       os.environ.get("SMTP_HOST",    cfg.get("smtp_host", "smtp.gmail.com")),
        "port":       int(os.environ.get("SMTP_PORT", cfg.get("smtp_port", 587))),
        "user":       os.environ.get("SMTP_USER",    cfg.get("smtp_user", "")),
        "passw":      os.environ.get("SMTP_PASS",    cfg.get("smtp_pass", "")),
        "from":       os.environ.get("SMTP_FROM",    cfg.get("smtp_from", "")),
        "resend_key": os.environ.get("RESEND_API_KEY", cfg.get("resend_api_key", "")),
    }

def _enviar_via_resend(api_key, de, para, assunto, html_body, reply_to=None):
    """Envia email usando a API HTTP do Resend (sem SMTP)."""
    data = {
        "from": de,
        "to": [para],
        "subject": assunto,
        "html": html_body
    }
    if reply_to:
        data["reply_to"] = [reply_to]
    payload = _json.dumps(data).encode("utf-8")
    req = urllib.request.Request(
        "https://api.resend.com/emails",
        data=payload,
        headers={
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
            "User-Agent": "python-urllib/sistema-boletos"
        },
        method="POST"
    )
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            return resp.status == 200
    except urllib.error.HTTPError as e:
        body = e.read().decode("utf-8", errors="ignore")
        raise Exception(f"HTTP {e.code}: {body}")

def _smtp_connect(smtp, timeout=15):
    """Conecta ao servidor SMTP usando SSL (porta 465) ou STARTTLS (demais portas)."""
    if smtp["port"] == 465:
        s = smtplib.SMTP_SSL(smtp["host"], smtp["port"], timeout=timeout)
    else:
        s = smtplib.SMTP(smtp["host"], smtp["port"], timeout=timeout)
        s.starttls()
    s.login(smtp["user"], smtp["passw"])
    return s

@app.route("/configurar_smtp", methods=["POST"])
@admin_required
def configurar_smtp():
    d = request.get_json() or {}
    cfg = ler_config()
    cfg["smtp_host"]      = d.get("host", "smtp.gmail.com")
    cfg["smtp_port"]      = int(d.get("port", 587))
    cfg["smtp_user"]      = d.get("user", "")
    cfg["smtp_pass"]      = d.get("passw", "")
    cfg["smtp_from"]      = d.get("from_addr", "")
    cfg["resend_api_key"] = d.get("resend_api_key", "")
    salvar_config(cfg)
    smtp = ler_smtp()
    # Se tiver chave Resend, testa via API
    if smtp["resend_key"]:
        try:
            _enviar_via_resend(
                smtp["resend_key"],
                "Funchal Imoveis <noreply@funchalimoveis.com.br>",
                smtp["user"] or "vithor.a.failde@gmail.com",
                "Teste de conexão — Sistema de Boletos",
                "<p>Conexão com Resend configurada com sucesso!</p>"
            )
            return jsonify({"ok": True})
        except Exception as ex:
            return jsonify({"ok": False, "erro": str(ex)})
    # Senão testa SMTP
    try:
        _smtp_connect(smtp, timeout=10)
        return jsonify({"ok": True})
    except Exception as ex:
        return jsonify({"ok": False, "erro": str(ex)})

@app.route("/get_smtp", methods=["GET"])
@login_required
def get_smtp():
    s = ler_smtp()
    configurado = bool(s["resend_key"]) or bool(s["user"] and s["passw"])
    return jsonify({"host": s["host"], "port": s["port"], "user": s["user"],
                    "from_addr": s["from"], "resend_key": "***" if s["resend_key"] else "",
                    "configurado": configurado})

@app.route("/emails_proprietarios", methods=["GET"])
@login_required
def get_emails_proprietarios():
    cfg = ler_config()
    return jsonify(cfg.get("emails_proprietarios", {}))

@app.route("/salvar_emails_proprietarios", methods=["POST"])
@login_required
def salvar_emails_proprietarios():
    emails = (request.get_json() or {}).get("emails", {})
    cfg = ler_config()
    cfg["emails_proprietarios"] = emails
    salvar_config(cfg)
    return jsonify({"ok": True})

def _gerar_html_email(proprietario, mes, rows, total, hoje):
    """Gera o corpo HTML do email do demonstrativo de repasse."""
    logo_path = BASE / "static" / "logo.png"
    logo_tag = ""
    if logo_path.exists():
        with open(logo_path, "rb") as f:
            b64 = base64.b64encode(f.read()).decode()
        logo_tag = f'<img src="data:image/png;base64,{b64}" style="max-width:100%;height:auto;display:block;margin-bottom:12px" alt="Funchal Imóveis">'

    def fmt(v):
        return f"R$ {float(v or 0):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

    unidades_html = ""
    for i, row in enumerate(rows):
        linhas = f'<tr><td style="padding:5px 10px;font-size:12pt">Aluguel bruto</td><td style="padding:5px 10px;text-align:right;font-size:12pt">{fmt(row["aluguel"])}</td></tr>'
        if row.get("taxa_imob"):
            linhas += f'<tr><td style="padding:5px 10px;font-size:11pt;font-style:italic;color:#555">(-) Taxa de administração {row.get("perc_imob","")}%</td><td style="padding:5px 10px;text-align:right;font-style:italic;color:#555">({fmt(row["taxa_imob"])})</td></tr>'
        if row.get("pa_ativo"):
            linhas += f'<tr><td style="padding:5px 10px;font-size:11pt;font-style:italic;color:#555">(-) Primeiro aluguel</td><td style="padding:5px 10px;text-align:right;font-style:italic;color:#555">({fmt(row["pa_val"])})</td></tr>'
        for d in (row.get("ded_prop_itens") or []):
            linhas += f'<tr><td style="padding:5px 10px;font-size:11pt;font-style:italic;color:#555">(-) {d["label"]} (prop.)</td><td style="padding:5px 10px;text-align:right;font-style:italic;color:#555">({fmt(d["val"])})</td></tr>'
        for d in (row.get("deds") or []):
            if d.get("sub"):
                linhas += f'<tr><td style="padding:5px 10px;font-size:11pt;font-style:italic;color:#555">(-) {d["desc"]}</td><td style="padding:5px 10px;text-align:right;font-style:italic;color:#555">({fmt(d["val"])})</td></tr>'
            else:
                linhas += f'<tr><td style="padding:5px 10px;font-size:11pt;font-style:italic;color:#336">(+) {d["desc"]}</td><td style="padding:5px 10px;text-align:right;font-style:italic;color:#336">{fmt(d["val"])}</td></tr>'
        linhas += f'<tr style="border-top:2px solid #000"><td style="padding:8px 10px;font-weight:bold;font-size:13pt">Valor líquido a repassar</td><td style="padding:8px 10px;text-align:right;font-weight:bold;font-size:13pt">{fmt(row["repasse"])}</td></tr>'

        sep = '<tr><td colspan="2" style="padding:0;height:14px;border:none"></td></tr>' if i > 0 else ''
        unidades_html += f'''
        {sep}
        <tr><td colspan="2" style="padding:6px 10px;font-weight:bold;font-size:11pt;text-transform:uppercase;border-bottom:1px solid #000;letter-spacing:.5px">
          Imóvel {i+1}: {row["endereco"]}<br>
          <span style="font-weight:normal;font-size:10pt;color:#555">Locatário: {row["locatario"]}{(" &nbsp;|&nbsp; " + row["tipo"]) if row.get("tipo") else ""}</span>
        </td></tr>
        {linhas}'''

    return f"""<!DOCTYPE html>
<html lang="pt-BR"><head><meta charset="UTF-8"></head>
<body style="font-family:'Times New Roman',Times,serif;color:#000;background:#fff;max-width:700px;margin:0 auto;padding:24px">
  {logo_tag}
  <div style="border-bottom:2px solid #000;padding-bottom:10px;margin-bottom:18px">
    <div style="font-size:16pt;font-weight:bold;text-transform:uppercase;letter-spacing:1px;text-align:center">Demonstrativo Mensal de Repasse</div>
    <div style="font-size:10pt;text-align:center;margin-top:3px">Competência: {mes}</div>
    <div style="display:flex;justify-content:space-between;margin-top:8px;font-size:9pt">
      <span>Data de emissão: {hoje}</span><span>Documento de uso interno</span>
    </div>
  </div>
  <table style="width:100%;border:1px solid #000;margin-bottom:18px;border-collapse:collapse">
    <tr><td style="padding:6px 10px;font-weight:bold;width:130px">Proprietário:</td>
        <td style="padding:6px 10px"><strong>{proprietario}</strong></td>
        <td style="padding:6px 10px;font-weight:bold;text-align:right;width:80px">Imóveis:</td>
        <td style="padding:6px 10px;width:30px">{len(rows)}</td></tr>
  </table>
  <table style="width:100%;border-collapse:collapse">
    {unidades_html}
  </table>
  <table style="width:100%;border:2px solid #000;margin-top:8px;border-collapse:collapse">
    <tr><td style="padding:12px 14px;font-weight:bold;font-size:13pt;text-transform:uppercase;letter-spacing:.5px">Total Líquido a Repassar</td>
        <td style="padding:12px 14px;text-align:right;font-weight:bold;font-size:16pt">{fmt(total)}</td></tr>
  </table>
  <div style="margin-top:24px;border-top:1px solid #000;padding-top:8px;font-size:9pt;color:#555;display:flex;justify-content:space-between">
    <span>Funchal Soluções Empresariais e Imobiliarias Ltda. — CRECI 21.360-J</span>
    <span>{hoje}</span>
  </div>
</body></html>"""

def gravar_log_envio(proprietario, email, mes, status, erro=""):
    """Acrescenta uma entrada no log de envios de informes."""
    import datetime as _dt
    try:
        log = []
        if LOG_ENVIOS_FILE.exists():
            with open(LOG_ENVIOS_FILE, encoding="utf-8") as f:
                log = json.load(f)
        log.insert(0, {
            "timestamp": _dt.datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
            "proprietario": proprietario,
            "email": email,
            "mes": mes,
            "status": status,   # "enviado" | "erro"
            "detalhe": erro,
        })
        log = log[:200]  # mantém só os últimos 200
        tmp = str(LOG_ENVIOS_FILE) + ".tmp"
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(log, f, ensure_ascii=False, indent=2)
        import os; os.replace(tmp, LOG_ENVIOS_FILE)
    except Exception:
        pass


@app.route("/api/log_envios", methods=["GET"])
@login_required
def api_log_envios():
    if not LOG_ENVIOS_FILE.exists():
        return jsonify([])
    with open(LOG_ENVIOS_FILE, encoding="utf-8") as f:
        return jsonify(json.load(f))


@app.route("/enviar_informe", methods=["POST"])
@login_required
def enviar_informe():
    d = request.get_json() or {}
    email_dest = (d.get("email_dest") or "").strip()
    proprietario = d.get("proprietario", "")
    mes   = d.get("mes", "")
    rows  = d.get("rows", [])
    total = d.get("total", 0)

    if not email_dest:
        return jsonify({"ok": False, "erro": "Email do destinatário não informado."})

    smtp = ler_smtp()
    if not smtp["resend_key"] and (not smtp["user"] or not smtp["passw"]):
        return jsonify({"ok": False, "erro": "Email não configurado. Configure em 'Configurar Email'."})

    hoje = __import__("datetime").date.today().strftime("%d/%m/%Y")
    html_body = _gerar_html_email(proprietario, mes, rows, total, hoje)
    assunto = f"Demonstrativo de Repasse — {proprietario} — {mes}"

    # Usa Resend se tiver chave
    if smtp["resend_key"]:
        remetente = "Funchal Imoveis <noreply@funchalimoveis.com.br>"
        reply_to  = smtp["user"] or None
        try:
            _enviar_via_resend(smtp["resend_key"], remetente, email_dest, assunto, html_body, reply_to=reply_to)
            gravar_log_envio(proprietario, email_dest, mes, "enviado")
            return jsonify({"ok": True})
        except Exception as ex:
            gravar_log_envio(proprietario, email_dest, mes, "erro", str(ex))
            return jsonify({"ok": False, "erro": str(ex)})

    # Senão usa SMTP
    remetente = smtp["from"] or smtp["user"]
    msg = MIMEMultipart("alternative")
    msg["Subject"] = assunto
    msg["From"]    = remetente
    msg["To"]      = email_dest
    msg.attach(MIMEText(html_body, "html", "utf-8"))
    try:
        with _smtp_connect(smtp, timeout=15) as s:
            s.sendmail(remetente, [email_dest], msg.as_string())
        gravar_log_envio(proprietario, email_dest, mes, "enviado")
        return jsonify({"ok": True})
    except Exception as ex:
        gravar_log_envio(proprietario, email_dest, mes, "erro", str(ex))
        return jsonify({"ok": False, "erro": str(ex)})


@app.route("/salvar_extras", methods=["POST"])
@login_required
def salvar_extras():
    historico = ler_historico()
    for row in (request.json or {}).get("locatarios", []):
        chave = row.get("chave") or norm(row.get("locatario", ""))
        if not chave:
            continue
        historico[chave] = {
            "iptu": row.get("iptu") or row.get("iptu_hist") or 0,
            "iptu_parcela": row.get("iptu_parcela") or row.get("iptu_parcela_hist") or "",
            "iptu_vaga": row.get("iptu_vaga") or row.get("iptu_vaga_hist") or 0,
            "iptu_vaga_parcela": row.get("iptu_vaga_parcela") or row.get("iptu_vaga_parcela_hist") or "",
            **{f"extras_fixo{n}_desc": row.get(f"extras_fixo{n}_desc") or row.get(f"extras_fixo{n}_desc_hist") or ""
               for n in ['', '2', '3', '4', '5']},
            **{f"extras_fixo{n}_val": row.get(f"extras_fixo{n}_val") or row.get(f"extras_fixo{n}_val_hist") or 0
               for n in ['', '2', '3', '4', '5']},
            "abono_desc": row.get("abono_desc") or row.get("abono_desc_hist") or "",
            "abono_val": row.get("abono_val") or row.get("abono_val_hist") or 0,
            "abono_parcela": row.get("abono_parcela") or row.get("abono_parcela_hist") or "",
            "seg_fianca": row.get("seg_fianca") or row.get("seg_fianca_hist") or 0,
            "seg_incendio": row.get("seg_incendio") or row.get("seg_incendio_hist") or 0,
            "cond_cota": float((row.get("cond_itens_rep") or {}).get("cota") or 0),
            **{f"ded{n}_desc": row.get(f"ded{n}_desc") or row.get(f"ded{n}_desc_hist") or ""
               for n in range(1, 11)},
            **{f"ded{n}_val":  row.get(f"ded{n}_val")  or row.get(f"ded{n}_val_hist")  or 0
               for n in range(1, 11)},
        }
    salvar_historico(historico)

    # Salva multa/juros no dimob_historico para uso na DIMOB
    mes_ref = (request.json or {}).get("mes", "")  # ex: "Junho/2026"
    if mes_ref:
        MESES_PT = {"jan":1,"fev":2,"mar":3,"abr":4,"mai":5,"jun":6,
                    "jul":7,"ago":8,"set":9,"out":10,"nov":11,"dez":12}
        partes = mes_ref.replace("/", " ").split()
        try:
            ano_ref  = int(partes[-1])
            mes_nome = partes[0].lower()[:3]
            mes_num  = MESES_PT.get(mes_nome)
        except Exception:
            ano_ref = mes_num = None

        if ano_ref and mes_num:
            hist_dimob = ler_dimob_historico()
            ano_str = str(ano_ref)
            if ano_str not in hist_dimob:
                hist_dimob[ano_str] = {}
            for row in (request.json or {}).get("locatarios", []):
                chave = row.get("chave") or norm(row.get("locatario", ""))
                if not chave:
                    continue
                multa = float(row.get("multa_atraso_val") or 0)
                juros = float(row.get("juros_mora_val") or 0)
                if multa or juros:
                    if chave not in hist_dimob[ano_str]:
                        hist_dimob[ano_str][chave] = {}
                    if "multa_juros" not in hist_dimob[ano_str][chave]:
                        hist_dimob[ano_str][chave]["multa_juros"] = {}
                    hist_dimob[ano_str][chave]["multa_juros"][str(mes_num)] = {
                        "multa": multa, "juros": juros
                    }
            salvar_dimob_historico(hist_dimob)

    return jsonify({"ok": True})


@app.route("/baixar_historico")
@login_required
def baixar_historico():
    """Faz download do historico.json para backup local do usuario."""
    if HISTORICO_FILE.exists():
        with open(HISTORICO_FILE, encoding="utf-8") as f:
            conteudo = f.read()
    else:
        conteudo = "{}"
    return Response(
        conteudo,
        mimetype="application/json",
        headers={"Content-Disposition": "attachment; filename=historico.json"}
    )

@app.route("/restaurar_historico", methods=["POST"])
@login_required
def restaurar_historico():
    """Restaura historico.json a partir de um arquivo enviado pelo usuario."""
    try:
        dados = request.get_json()
        if not isinstance(dados, dict):
            return jsonify({"ok": False, "erro": "Formato invalido — envie um JSON valido"})
        salvar_historico(dados)
        return jsonify({"ok": True, "registros": len(dados)})
    except Exception as e:
        return jsonify({"ok": False, "erro": str(e)})

@app.route("/restaurar_dimob_historico", methods=["POST"])
@login_required
def restaurar_dimob_historico():
    """Restaura dimob_historico.json a partir de um arquivo enviado pelo usuario."""
    try:
        dados = request.get_json()
        if not isinstance(dados, dict):
            return jsonify({"ok": False, "erro": "Formato invalido — envie um JSON valido"})
        salvar_dimob_historico(dados)
        total = sum(len(v) for v in dados.values() if isinstance(v, dict))
        return jsonify({"ok": True, "registros": total})
    except Exception as e:
        return jsonify({"ok": False, "erro": str(e)})

@app.route("/exportar", methods=["POST"])
@login_required
def exportar():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    dados = (request.json or {}).get("locatarios", [])
    mes = (request.json or {}).get("mes", "")

    wb = Workbook()
    ws = wb.active
    ws.title = "Boletos"
    cols = ["Locatario", "Endereco", "Tipo", "Proprietario",
            "Aluguel", "Cond Repasse", "Fundo Reserva*",
            "IPTU Apto", "Parc IPTU", "IPTU Vaga", "Parc IPTU Vaga",
            "Seg Fianca", "Seg Incendio",
            "Outros Desc", "Outros Valor",
            "Outros Fixo Desc", "Outros Fixo Valor",
            "Abono Desc", "Abono Valor",
            "TOTAL",
            "Taxa Imob (%)", "Taxa Imob (R$)", "Deducoes Prop.",
            "Ded. Manuais Desc", "Ded. Manuais (R$)",
            "REPASSE PROPRIETARIO"]
    # col 26 = REPASSE PROPRIETARIO (1-based)

    hf = PatternFill("solid", start_color="1E3A5F")
    hfont = Font(color="FFFFFF", bold=True, size=10)
    bdr = Border(bottom=Side(style="thin", color="E2E8F0"),
                 right=Side(style="thin", color="E2E8F0"))

    hf_green = PatternFill("solid", start_color="15803D")
    for c, h in enumerate(cols, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = hfont
        cell.fill = hf_green if c == 26 else hf
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    for i, row in enumerate(dados, 2):
        al = row.get("aluguel") or 0
        co = row.get("cond_repasse") or 0
        fnr = row.get("cond_nao_repasse") or 0
        # Frontend já inicializa os valores efetivos (hist aplicado), sem fallback necessário
        ip  = float(row.get("iptu")            or 0)
        ipv = float(row.get("iptu_vaga")       or 0)
        sf  = float(row.get("seg_fianca")      or 0)
        si  = float(row.get("seg_incendio")    or 0)
        ev  = float(row.get("extras_val")        or 0)
        ev2 = float(row.get("extras2_val")       or 0)
        ev3 = float(row.get("extras3_val")       or 0)
        ef  = float(row.get("extras_fixo_val")   or 0)
        ef2 = float(row.get("extras_fixo2_val")  or 0)
        ef3 = float(row.get("extras_fixo3_val")  or 0)
        ef4 = float(row.get("extras_fixo4_val")  or 0)
        ef5 = float(row.get("extras_fixo5_val")  or 0)
        ab  = float(row.get("abono_val")         or 0)
        total = al + co + ip + ipv + sf + si + ev + ev2 + ev3 + ef + ef2 + ef3 + ef4 + ef5 - ab
        ip_parc = row.get("iptu_parcela") or row.get("iptu_parcela_hist") or ""
        ipv_parc = row.get("iptu_vaga_parcela") or row.get("iptu_vaga_parcela_hist") or ""

        taxa_imob_p = float(row.get("percentual_imob") or 0)
        taxa_imob_r = float(row.get("taxa_imob") or 0)
        ded_prop    = float(row.get("cond_nao_repasse") or 0)
        def _ded_signed(n):
            val = safe_float(row.get(f"ded{n}_val"))
            sub = row.get(f"ded{n}_subtrair", True)
            return val if (sub is None or sub) else -val
        ded_man_val = sum(_ded_signed(n) for n in range(1, 11))
        ded_man_desc = " | ".join(
            (row.get(f"ded{n}_desc") or "").strip()
            for n in range(1, 11)
            if safe_float(row.get(f"ded{n}_val"))
        )
        rep_prop    = al - taxa_imob_r - ded_prop - ded_man_val

        vals = [row.get("locatario", ""), row.get("endereco", ""), row.get("tipo", ""),
                row.get("proprietario", ""), al, co, fnr,
                ip, ip_parc,
                ipv, ipv_parc,
                sf, si,
                row.get("extras_desc", ""), ev,
                row.get("extras2_desc", ""), ev2,
                row.get("extras3_desc", ""), ev3,
                row.get("extras_fixo_desc")  or row.get("extras_fixo_desc_hist")  or "", ef,
                row.get("extras_fixo2_desc") or row.get("extras_fixo2_desc_hist") or "", ef2,
                row.get("extras_fixo3_desc") or row.get("extras_fixo3_desc_hist") or "", ef3,
                row.get("extras_fixo4_desc") or row.get("extras_fixo4_desc_hist") or "", ef4,
                row.get("extras_fixo5_desc") or row.get("extras_fixo5_desc_hist") or "", ef5,
                row.get("abono_desc") or row.get("abono_desc_hist") or "", ab,
                total,
                taxa_imob_p, taxa_imob_r, ded_prop,
                ded_man_desc, ded_man_val,
                rep_prop]

        # colunas com formato moeda (1-based): 5=Aluguel, 6=Cond, 7=FundoRes,
        # 8=IPTU, 10=IPTUVaga, 12=SegFianca, 13=SegIncendio,
        # 15=OutrosVal, 17=FixoVal, 19=AbVal, 20=TOTAL,
        # 22=TaxaImobR$, 23=DeducoesProp, 25=DedManVal, 26=RepasseProp
        moeda_cols = {5, 6, 7, 8, 10, 12, 13, 15, 17, 19, 20, 22, 23, 25, 26}
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=i, column=c, value=v)
            cell.border = bdr
            if c in moeda_cols:
                cell.number_format = 'R$ #,##0.00'
            if c == 20:
                cell.font = Font(bold=True)
            if c == 26:
                cell.font = Font(bold=True, color="15803D")

    for col in ws.columns:
        w = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(w + 3, 44)

    ws.cell(row=len(dados)+3, column=7,
            value="* Fundo de reserva: encargo do proprietario, nao incluido no total do locatario")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    nome = f"boletos_{mes.replace('/', '_') if mes else 'export'}.xlsx"
    return Response(buf.read(),
                    mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f"attachment; filename={nome}"})


# ─── Rotas: Reajuste de Aluguel ───────────────────────────────────────────────

def _excel_path_contratos():
    """Retorna o caminho da planilha de contratos usada por AMBOS os sistemas
    (boletos e reajustes). Prioridade: contratos.xlsx → primeiro .xlsx encontrado.
    """
    p = UPLOAD_DIR / 'contratos.xlsx'
    if p.exists():
        return p
    fallback = next(UPLOAD_DIR.glob('*.xlsx'), None)
    return fallback  # pode ser None se não houver nenhum arquivo


@app.route('/reajustes')
@login_required
def reajustes_page():
    return render_template('reajustes.html', tem_excel=False)


@app.route('/api/calcular_reajustes', methods=['POST'])
@login_required
def api_calcular_reajustes():
    """Lê a planilha enviada pelo usuário, consulta BACEN e devolve JSON com
    todos os contratos + acumulado 12m + novo aluguel sugerido.
    """
    # 1. Requer upload explícito — nunca usa planilha de sessão anterior
    excel_file = request.files.get('excel')
    if not excel_file or not excel_file.filename.lower().endswith(('.xlsx', '.xls')):
        return jsonify({'ok': False, 'erro': 'Selecione a planilha de contratos antes de calcular.'})
    excel_path = UPLOAD_DIR / 'contratos.xlsx'
    excel_file.save(str(excel_path))

    # 2. Ler contratos
    try:
        contratos = ler_excel_reajustes(excel_path)
    except Exception as e:
        return jsonify({'ok': False, 'erro': f'Erro ao ler planilha: {e}'})

    if not contratos:
        return jsonify({'ok': False, 'erro': 'Nenhum contrato válido encontrado na planilha.'})

    # 3. Buscar dados de índices — uma chamada por índice, não por contrato.
    #    Estratégia: número-índice (precisão máxima) para contratos ESTE_MES,
    #    variação mensal composta (BACEN SGS) para contratos FUTURO (dados parciais).
    indices_necessarios = set(c['indice_norm'] for c in contratos)
    tem_este_mes = any(c['status'] == 'ESTE_MES' for c in contratos)

    historicos_mensal  = {}   # % mensal — IGPM usa série 28655 (alta precisão), demais série padrão
    historicos_indice  = {}   # número-índice — para ESTE_MES IPCA/INPC (precisão máxima)
    erros_bacen        = {}

    for idx in indices_necessarios:
        # ── % mensal (BACEN SGS) — sempre busca; IGPM usa série 28655 alta precisão ──
        sid_mensal = BACEN_SERIES.get(idx)
        if sid_mensal:
            hist_m, err_m = buscar_historico_bacen(sid_mensal, anos=5)
            historicos_mensal[idx] = hist_m
            if err_m:
                erros_bacen[idx] = err_m

        # ── IPCA/INPC: número-índice BACEN (séries 1737/1617) — só se houver ESTE_MES ──
        if not tem_este_mes or idx == 'IGPM':
            continue
        sid_indice = BACEN_SERIES_INDICE.get(idx)
        if sid_indice:
            hist_i, err_i = buscar_historico_bacen(sid_indice, anos=3, tentativas=2, timeout=15)
            if hist_i:
                historicos_indice[idx] = hist_i
            # falha silenciosa — fallback automático para série mensal no cálculo

    # 4a. Para contratos ESTE_MES com IPCA e aniversário antes do dia 15:
    #     o IPCA do mês de aniversário pode não estar publicado ainda no BACEN.
    #     Nesse caso, replica o % do mês anterior no lugar do mês ausente.
    for c in contratos:
        if c['status'] != 'ESTE_MES' or c.get('indice_norm') != 'IPCA':
            continue
        dr = date.fromisoformat(c['data_reajuste_iso'])
        if dr.day >= 15:
            continue
        key_aniv = f"{dr.year}-{dr.month:02d}"
        m_p = dr.month - 1 if dr.month > 1 else 12
        a_p = dr.year if dr.month > 1 else dr.year - 1
        key_prev = f"{a_p}-{m_p:02d}"
        # Série mensal 433 — usada pelo fallback e pelo acumulado FUTURO
        hm = historicos_mensal.get('IPCA', {})
        if key_aniv not in hm and key_prev in hm:
            hm[key_aniv] = hm[key_prev]
        # Série número-índice 1737 — estima Index(M) replicando a variação de M-1
        hi = historicos_indice.get('IPCA', {})
        if key_aniv not in hi and key_prev in hi:
            m_p2 = m_p - 1 if m_p > 1 else 12
            a_p2 = a_p if m_p > 1 else a_p - 1
            key_prev2 = f"{a_p2}-{m_p2:02d}"
            if key_prev2 in hi:
                hi[key_aniv] = hi[key_prev] ** 2 / hi[key_prev2]
            else:
                hi[key_aniv] = hi[key_prev]

    # 4. Calcular acumulado e novo aluguel — apenas para contratos ESTE_MES e FUTURO
    #    Contratos OK (já reajustados) não precisam de cálculo: o Excel já está correto
    hoje = date.today()
    for c in contratos:
        if c['status'] == 'OK' or not c.get('vigencia_ok', True):
            c['acumulado_pct'] = None
            c['meses_base']    = 0
            c['novo_aluguel']  = None
            c['diferenca']     = None
            c['aplicavel']     = False
            continue

        metodo = 'mensal'
        try:
            data_rej = date.fromisoformat(c['data_reajuste_iso'])
            idx = c['indice_norm']
            if c['status'] == 'ESTE_MES':
                if idx == 'IGPM':
                    # Série 28655 (alta precisão) já está em historicos_mensal['IGPM']
                    acum, meses = calcular_acumulado_12m(
                        historicos_mensal.get('IGPM', {}), data_rej)
                elif idx in historicos_indice:
                    # IPCA / INPC — BACEN 1737/1617: número-índice real (base dez/1993=100)
                    acum, meses = calcular_por_numero_indice(
                        historicos_indice[idx], data_rej)
                    if acum is None:
                        acum, meses = calcular_acumulado_12m(
                            historicos_mensal.get(idx, {}), data_rej)
                else:
                    acum, meses = calcular_acumulado_12m(
                        historicos_mensal.get(idx, {}), data_rej)
            else:
                # FUTURO — % mensal composto (dados parciais)
                acum, meses = calcular_acumulado_12m(
                    historicos_mensal.get(idx, {}), data_rej)
        except Exception:
            acum, meses = None, 0
        c['metodo_calculo'] = metodo

        c['acumulado_pct'] = acum
        c['meses_base']    = meses
        if acum is not None:
            novo = round(c['aluguel'] * (1 + acum / 100), 2)
            c['novo_aluguel'] = novo
            c['diferenca']    = round(novo - c['aluguel'], 2)
        else:
            c['novo_aluguel'] = None
            c['diferenca']    = None

        # Aplicavel apenas no mes de aplicacao (ESTE_MES)
        c['aplicavel'] = (c['status'] == 'ESTE_MES') and c['novo_aluguel'] is not None

    # 5. Estatísticas
    total       = len(contratos)
    este_mes    = sum(1 for c in contratos if c['status'] == 'ESTE_MES')
    futuros     = sum(1 for c in contratos if c['status'] == 'FUTURO')
    ok_count    = total - este_mes - futuros
    com_calculo = sum(1 for c in contratos if c['acumulado_pct'] is not None)
    aplicaveis  = sum(1 for c in contratos if c.get('aplicavel'))

    return jsonify({
        'ok': True,
        'contratos': contratos,
        'resumo': {
            'total': total,
            'este_mes': este_mes,
            'futuros': futuros,
            'ok': ok_count,
            'com_calculo': com_calculo,
            'aplicaveis': aplicaveis,
            'erros_bacen': erros_bacen,
            'data_consulta': hoje.strftime('%d/%m/%Y'),
        }
    })


@app.route('/api/aplicar_reajustes', methods=['POST'])
@login_required
def api_aplicar_reajustes():
    """Recebe lista de contratos aprovados pelo usuário, escreve os novos
    aluguéis e as novas datas de aniversário diretamente na planilha salva.
    A data de aniversário avança 1 ano automaticamente.
    """
    dados = request.get_json(silent=True) or {}
    contratos_aplicar = dados.get('contratos', [])
    if not contratos_aplicar:
        return jsonify({'ok': False, 'erro': 'Nenhum contrato enviado para aplicar.'})

    excel_path = _excel_path_contratos()
    if not excel_path:
        return jsonify({'ok': False, 'erro': 'Planilha não encontrada no servidor.'})

    try:
        n, erros = aplicar_reajustes_excel(excel_path, contratos_aplicar)
    except Exception as e:
        return jsonify({'ok': False, 'erro': f'Erro ao atualizar planilha: {e}'})

    return jsonify({
        'ok': True,
        'atualizados': n,
        'erros': erros,
        'msg': (f'{n} contrato(s) atualizado(s) na planilha. '
                'O sistema de boletos já usará os novos valores no próximo processamento.')
    })


@app.route('/api/debug_indices')
@login_required
def api_debug_indices():
    """Diagnóstico: IGPM (série 28655) e IPCA (série 1737) para aniversário abril/2026."""
    aniv = date(2026, 4, 30)
    key_ini = f"{aniv.year - 1}-{aniv.month:02d}"
    key_fim = f"{aniv.year}-{aniv.month:02d}"

    # ── IGPM série 28655 (alta precisão, produto 13m) ────────────────────────
    hist_igpm, err_igpm = buscar_historico_bacen(28655, anos=3)
    acum_igpm, meses_igpm = calcular_acumulado_12m(hist_igpm, aniv) if hist_igpm else (None, 0)

    # ── IPCA série 1737 (número-índice, razão exata) ─────────────────────────
    hist_ipca_idx, err_ipca_idx = buscar_historico_bacen(1737, anos=3, tentativas=2, timeout=15)
    acum_ipca_idx, _ = calcular_por_numero_indice(hist_ipca_idx, aniv) if hist_ipca_idx else (None, 0)

    # ── IPCA série 433 (mensal %, fallback) ──────────────────────────────────
    hist_ipca_mensal, err_ipca_m = buscar_historico_bacen(433, anos=3)
    acum_ipca_mensal, meses_ipca_m = calcular_acumulado_12m(hist_ipca_mensal, aniv) if hist_ipca_mensal else (None, 0)

    return jsonify({
        'IGPM': {
            'serie': 28655,
            'metodo': 'produto 13 meses (alta precisão)',
            'acumulado': acum_igpm,
            'meses': meses_igpm,
            'erro': err_igpm,
        },
        'IPCA': {
            'numero_indice_1737': {
                'valor_ini': hist_ipca_idx.get(key_ini) if hist_ipca_idx else None,
                'valor_fim': hist_ipca_idx.get(key_fim) if hist_ipca_idx else None,
                'acumulado': acum_ipca_idx,
                'erro': err_ipca_idx,
            },
            'mensal_433_fallback': {
                'acumulado_13m': acum_ipca_mensal,
                'meses': meses_ipca_m,
                'erro': err_ipca_m,
            },
        },
    })

@app.route('/api/baixar_contratos')
@login_required
def api_baixar_contratos():
    """Faz o download da planilha de contratos atualizada (com reajustes aplicados)."""
    from flask import send_file
    excel_path = _excel_path_contratos()
    if not excel_path or not excel_path.exists():
        return jsonify({'ok': False, 'erro': 'Planilha não encontrada.'}), 404
    return send_file(
        str(excel_path),
        as_attachment=True,
        download_name='contratos_locacao_editado.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


# ─── Rotas: DIMOB ─────────────────────────────────────────────────────────────

@app.route('/dimob')
@login_required
def dimob_page():
    return render_template('dimob.html')


@app.route('/api/dimob_calcular', methods=['POST'])
@login_required
def api_dimob_calcular():
    """Calcula informes DIMOB para todos os contratos do ano de referência.
    Aceita multipart/form-data com campo 'ano' e opcionalmente 'excel'.
    """
    ano_str    = request.form.get('ano', str(date.today().year - 1))
    excel_file = request.files.get('excel')

    try:
        ano = int(ano_str)
    except ValueError:
        return jsonify({'ok': False, 'erro': 'Ano inválido.'})

    excel_path = _excel_path_contratos()
    if excel_file and excel_file.filename.lower().endswith(('.xlsx', '.xls')):
        excel_path = UPLOAD_DIR / 'contratos.xlsx'
        excel_file.save(str(excel_path))

    if not excel_path or not excel_path.exists():
        return jsonify({'ok': False, 'erro': 'Planilha não encontrada. Faça o upload da planilha de contratos.'})

    try:
        contratos = ler_excel_dimob(excel_path)
    except Exception as e:
        return jsonify({'ok': False, 'erro': f'Erro ao ler planilha: {e}'})

    if not contratos:
        return jsonify({'ok': False, 'erro': 'Nenhum contrato válido encontrado na planilha.'})

    historico_dimob = ler_dimob_historico()
    resultado = []

    for c in contratos:
        meses, tem_historico = calcular_meses_dimob(c, ano, historico_dimob)
        total    = round(sum(v for v in meses if v is not None), 2)
        comissao = round(total * c['percentual_imob'] / 100, 2) if c['percentual_imob'] else 0.0

        # Busca aluguel_antigo para exibição correta na coluna da tabela
        chave_dim      = norm(c['locatario'])
        hist_c         = historico_dimob.get(str(ano), {}).get(chave_dim, {})
        aluguel_antigo = hist_c.get('aluguel_antigo') if tem_historico else None

        d_ini = c.get('data_inicio')
        d_fim = c.get('data_fim')
        resultado.append({
            'locatario':        c['locatario'],
            'proprietario':     c['proprietario'],
            'endereco':         c['endereco'],
            'cpf_proprietario': c['cpf_proprietario'],
            'cpf_locatario':    c['cpf_locatario'],
            'aluguel_atual':    c['aluguel'],
            'aluguel_antigo':   aluguel_antigo,
            'mes_aplicacao':    c['mes_aplicacao'],
            'meses':            meses,
            'total':            total,
            'comissao':         comissao,
            'tem_historico':    tem_historico,
            'percentual_imob':  c['percentual_imob'],
            'data_inicio_str':  d_ini.strftime('%d/%m/%Y') if d_ini else None,
            'data_fim_str':     d_fim.strftime('%d/%m/%Y') if d_fim else None,
        })

    return jsonify({
        'ok':          True,
        'ano':         ano,
        'contratos':   resultado,
        'total_geral': round(sum(r['total'] for r in resultado), 2),
    })


@app.route('/api/dimob_salvar_historico', methods=['POST'])
@login_required
def api_dimob_salvar_historico():
    """Salva aluguéis anteriores informados manualmente no histórico DIMOB."""
    dados = request.get_json(silent=True) or {}
    ano   = str(dados.get('ano', date.today().year))
    contratos = dados.get('contratos', [])

    if not contratos:
        return jsonify({'ok': False, 'erro': 'Nenhum dado enviado.'})

    historico = ler_dimob_historico()
    if ano not in historico:
        historico[ano] = {}

    for c in contratos:
        locatario     = c.get('locatario', '')
        aluguel_antigo = c.get('aluguel_antigo')
        if not locatario or not aluguel_antigo:
            continue
        chave = norm(locatario)
        historico[ano][chave] = {
            'aluguel_antigo': float(aluguel_antigo),
            'aluguel_novo':   float(c.get('aluguel_novo', 0)),
            'locatario':      locatario,
            'mes_aplicacao':  c.get('mes_aplicacao'),
        }

    salvar_dimob_historico(historico)
    return jsonify({'ok': True, 'salvos': len(contratos)})


@app.route('/api/dimob_exportar', methods=['POST'])
@login_required
def api_dimob_exportar():
    """Exporta Informes de Pagamentos em Excel — um sheet por proprietário."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    dados     = request.get_json(silent=True) or {}
    ano       = dados.get('ano', date.today().year - 1)
    contratos = dados.get('contratos', [])

    if not contratos:
        return jsonify({'ok': False, 'erro': 'Sem dados para exportar.'}), 400

    wb = Workbook()
    wb.remove(wb.active)

    cor_header  = PatternFill("solid", start_color="2A4A52")
    cor_alt     = PatternFill("solid", start_color="EFF5F7")
    fonte_neg   = Font(bold=True)
    fonte_hdr   = Font(color="FFFFFF", bold=True, size=10)
    fonte_tot   = Font(bold=True, size=11)
    fonte_tit   = Font(bold=True, size=13)
    borda_fina  = Border(
        left=Side(style='thin', color='B8CECE'),
        right=Side(style='thin', color='B8CECE'),
        top=Side(style='thin', color='B8CECE'),
        bottom=Side(style='thin', color='B8CECE'),
    )
    centro = Alignment(horizontal='center', vertical='center')
    direita = Alignment(horizontal='right', vertical='center')
    esquerda = Alignment(horizontal='left', vertical='center')

    por_prop = {}
    for c in contratos:
        prop = c.get('proprietario') or 'SEM PROPRIETÁRIO'
        por_prop.setdefault(prop, []).append(c)

    sheet_names_usados = {}
    for prop_nome, cs in por_prop.items():
        base = re.sub(r'[\\/*?:\[\]]', '', prop_nome)[:28].strip()
        if not base:
            base = 'SEM_NOME'
        contagem = sheet_names_usados.get(base, 0)
        sheet_names_usados[base] = contagem + 1
        sheet_name = base if contagem == 0 else f"{base[:25]}_{contagem}"
        ws = wb.create_sheet(title=sheet_name)
        ws.column_dimensions['A'].width = 22
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 16
        ws.column_dimensions['F'].width = 18

        linha_atual = 1
        total_prop  = 0.0

        for c in cs:
            # ── Título ──
            ws.merge_cells(f'A{linha_atual}:F{linha_atual}')
            cel = ws.cell(row=linha_atual, column=1, value=f'INFORME DE PAGAMENTOS {ano}')
            cel.font = fonte_tit
            cel.alignment = centro
            ws.row_dimensions[linha_atual].height = 24
            linha_atual += 1

            # ── Emissão ──
            ws.merge_cells(f'A{linha_atual}:F{linha_atual}')
            cel = ws.cell(row=linha_atual, column=1,
                          value=f'Emissão em {date.today().strftime("%d/%m/%Y")}')
            cel.alignment = centro
            ws.row_dimensions[linha_atual].height = 16
            linha_atual += 1

            # ── Proprietário ──
            ws.cell(row=linha_atual, column=1, value='Proprietário:').font = fonte_neg
            ws.cell(row=linha_atual, column=2, value=c['proprietario'])
            ws.cell(row=linha_atual, column=5, value='CNPJ / CPF :').font = fonte_neg
            ws.cell(row=linha_atual, column=6, value=c['cpf_proprietario'] or '')
            ws.row_dimensions[linha_atual].height = 15
            linha_atual += 1

            # ── Imóvel ──
            ws.cell(row=linha_atual, column=1, value='Imóvel:').font = fonte_neg
            ws.merge_cells(f'B{linha_atual}:F{linha_atual}')
            ws.cell(row=linha_atual, column=2, value=c['endereco'])
            ws.row_dimensions[linha_atual].height = 15
            linha_atual += 1

            # ── Locatário ──
            ws.cell(row=linha_atual, column=1, value='Locatário:').font = fonte_neg
            ws.merge_cells(f'B{linha_atual}:D{linha_atual}')
            ws.cell(row=linha_atual, column=2, value=c['locatario'])
            ws.cell(row=linha_atual, column=5, value='CPF').font = fonte_neg
            ws.cell(row=linha_atual, column=6, value=c['cpf_locatario'] or '')
            ws.row_dimensions[linha_atual].height = 15
            linha_atual += 1

            # ── Cabeçalho da tabela ──
            for col_i, titulo in enumerate(['Mês Pagamento', 'ALUGUEL', '', '', '', 'Valor Líquido'], 1):
                cel = ws.cell(row=linha_atual, column=col_i, value=titulo)
                cel.fill = cor_header
                cel.font = fonte_hdr
                cel.alignment = centro
                cel.border = borda_fina
            ws.row_dimensions[linha_atual].height = 18
            linha_atual += 1

            # ── Meses ──
            total_imovel = 0.0
            for i, (nome_mes, val) in enumerate(zip(MESES_NOMES, c['meses'])):
                fill = cor_alt if i % 2 == 0 else PatternFill("solid", start_color="FFFFFF")
                cel_mes = ws.cell(row=linha_atual, column=1, value=nome_mes)
                cel_mes.font = fonte_neg
                cel_mes.fill = fill
                cel_mes.border = borda_fina

                if val is not None:
                    total_imovel += val
                    cel_val = ws.cell(row=linha_atual, column=2, value=val)
                    cel_val.number_format = '#,##0.00'
                    cel_val.fill = fill
                    cel_val.border = borda_fina
                    cel_liq = ws.cell(row=linha_atual, column=6, value=val)
                    cel_liq.number_format = '#,##0.00'
                    cel_liq.fill = fill
                    cel_liq.border = borda_fina
                else:
                    ws.cell(row=linha_atual, column=2, value='-').border = borda_fina
                    ws.cell(row=linha_atual, column=6, value='-').border = borda_fina

                ws.row_dimensions[linha_atual].height = 15
                linha_atual += 1

            total_prop += total_imovel

            # ── Total do imóvel ──
            cel_lbl = ws.cell(row=linha_atual, column=1, value='Total do Imóvel')
            cel_lbl.font = fonte_tot
            cel_lbl.border = borda_fina
            cel_tv = ws.cell(row=linha_atual, column=2, value=round(total_imovel, 2))
            cel_tv.font = fonte_tot
            cel_tv.number_format = '#,##0.00'
            cel_tv.border = borda_fina
            cel_tl = ws.cell(row=linha_atual, column=6, value=round(total_imovel, 2))
            cel_tl.font = fonte_tot
            cel_tl.number_format = '#,##0.00'
            cel_tl.border = borda_fina
            ws.row_dimensions[linha_atual].height = 18
            linha_atual += 3  # espaço entre imóveis

        # ── Total do Proprietário ──
        ws.cell(row=linha_atual, column=1, value='Total do Proprietário').font = fonte_tot
        cel_tp = ws.cell(row=linha_atual, column=2, value=round(total_prop, 2))
        cel_tp.font = fonte_tot
        cel_tp.number_format = '#,##0.00'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    nome = f"informes_dimob_{ano}.xlsx"
    return Response(
        buf.read(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={"Content-Disposition": f"attachment; filename={nome}"}
    )


# ─── Envio de Boletos aos Locatários ─────────────────────────────────────────

PROMPT_NOME_LOCATARIO = """Analise este boleto de locação/aluguel brasileiro.
Extraia o nome do locatário/inquilino e o mês de referência.

Retorne APENAS JSON válido:
{
  "locatario": "nome completo do locatário",
  "mes_referencia": "Mês/Ano ex: Abril/2026"
}

Instruções:
- O locatário é quem paga o aluguel (não o proprietário do imóvel)
- Procure campos: "Locatário:", "Inquilino:", "Sacado:", "Nome do Locatário:", "Pagador:"
- Para o mês procure: "Referente a:", "Competência:", "Vencimento:", "Mês/Ano"
- Retorne null se não encontrar o campo"""


@app.route("/envio_boletos")
@login_required
def envio_boletos():
    smtp = ler_smtp()
    config = ler_config()
    tem_smtp = bool(smtp["resend_key"]) or bool(smtp["user"] and smtp["passw"])
    return render_template("envio_boletos.html",
                           tem_smtp=tem_smtp,
                           tem_api_key=bool(config.get("api_key")),
                           is_admin=is_admin())


@app.route("/api/locatarios_emails", methods=["GET"])
@login_required
def get_locatarios_emails():
    emails_salvos = ler_locatarios_emails()
    resultado = dict(emails_salvos)

    # Complementa com locatários do arquivo de nomes (salvo ao processar boletos)
    nomes_path = DATA_DIR / "locatarios_nomes.json"
    if nomes_path.exists():
        try:
            with open(nomes_path, encoding="utf-8") as _f:
                nomes_salvos = json.load(_f)
            for chave, nome in nomes_salvos.items():
                if chave not in resultado:
                    resultado[chave] = {"locatario": nome, "email": ""}
        except Exception:
            pass

    # Fallback: tenta ler direto da planilha Excel (se existir no upload dir)
    if not resultado:
        excel_path = UPLOAD_DIR / "contratos.xlsx"
        if excel_path.exists():
            try:
                contratos, _ = ler_excel(excel_path)
                for _, ct in contratos.items():
                    nome = ct["locatario"]
                    chave = norm(nome)
                    if chave not in resultado:
                        resultado[chave] = {"locatario": nome, "email": ""}
            except Exception:
                pass

    return jsonify(resultado)


@app.route("/api/locatarios_emails/bulk", methods=["POST"])
@login_required
def bulk_locatarios_emails():
    """Salva todos os e-mails de locatários de uma vez."""
    emails_novos = (request.get_json() or {}).get("emails", {})
    emails = ler_locatarios_emails()
    for chave, info in emails_novos.items():
        emails[chave] = info
    salvar_locatarios_emails(emails)
    return jsonify({"ok": True})


@app.route("/api/locatarios_emails", methods=["POST"])
@login_required
def post_locatario_email():
    d = request.get_json() or {}
    nome = (d.get("locatario") or "").strip()
    email_loc = (d.get("email") or "").strip()
    chave_antiga = (d.get("chave_antiga") or "").strip()
    if not nome:
        return jsonify({"ok": False, "erro": "Nome obrigatório"})
    emails = ler_locatarios_emails()
    # Se editando (chave_antiga diferente da nova), remove a antiga
    nova_chave = norm(nome)
    if chave_antiga and chave_antiga != nova_chave and chave_antiga in emails:
        del emails[chave_antiga]
    emails[nova_chave] = {"locatario": nome, "email": email_loc}
    salvar_locatarios_emails(emails)
    return jsonify({"ok": True, "chave": nova_chave})


@app.route("/api/locatarios_emails/<chave>", methods=["DELETE"])
@login_required
def delete_locatario_email(chave):
    emails = ler_locatarios_emails()
    if chave in emails:
        del emails[chave]
        salvar_locatarios_emails(emails)
    return jsonify({"ok": True})


@app.route("/envio_boletos/processar", methods=["POST"])
@login_required
def processar_boletos_locatarios():
    """Identifica o locatário de cada boleto pelo nome do arquivo (sem usar IA)."""
    pdf_files = request.files.getlist("boletos")
    if not pdf_files:
        return jsonify({"ok": False, "erro": "Nenhum PDF enviado"})

    emails_db = ler_locatarios_emails()

    # Salva arquivos no upload dir
    saved = []
    for f in pdf_files:
        if f.filename.lower().endswith(".pdf"):
            safe_name = Path(f.filename).name
            dest = UPLOAD_DIR / ("loc_" + safe_name)
            f.save(str(dest))
            saved.append(safe_name)

    def _match_nome_arquivo(nome_arq):
        """Extrai palavras do nome do arquivo e tenta casar com o banco de emails."""
        stem = Path(nome_arq).stem  # remove extensão
        # Normaliza: remove underscores, hifens e palavras genéricas como "boleto"
        stem = re.sub(r'(?i)boleto[_\s-]*', '', stem)
        stem = re.sub(r'[_\-]', ' ', stem).strip()
        palavras_arq = norm_palavras(stem)

        chave_match = None
        email_match = None
        nome_match = None
        melhor_score = 0

        # Match exato primeiro
        chave_exato = norm(stem)
        if chave_exato in emails_db:
            return chave_exato, emails_db[chave_exato].get("locatario", stem), emails_db[chave_exato].get("email", ""), 100

        # Match por palavras
        for chave, info in emails_db.items():
            palavras_db = norm_palavras(info.get("locatario", ""))
            s = score_nome_arquivo(palavras_arq, palavras_db)
            sc = int(s * 100)
            if sc > melhor_score and sc >= 50:
                melhor_score = sc
                chave_match = chave
                email_match = info.get("email", "")
                nome_match = info.get("locatario", "")

        return chave_match, nome_match, email_match, melhor_score

    def gerar():
        yield f"data: {json.dumps({'tipo': 'inicio', 'total': len(saved)})}\n\n"
        resultados = []

        for i, orig_name in enumerate(saved):
            yield f"data: {json.dumps({'tipo': 'progresso', 'atual': i+1, 'total': len(saved), 'arquivo': orig_name})}\n\n"

            chave_match, nome_match, email_match, melhor_score = _match_nome_arquivo(orig_name)
            # Nome extraído do arquivo (para exibir na tabela)
            stem_display = re.sub(r'(?i)boleto[_\s-]*', '', Path(orig_name).stem)
            stem_display = re.sub(r'[_\-]', ' ', stem_display).strip()

            resultado = {
                "arquivo": orig_name,
                "arquivo_salvo": "loc_" + orig_name,
                "locatario_pdf": stem_display,
                "mes_referencia": "",
                "chave_match": chave_match,
                "locatario_match": nome_match,
                "email": email_match,
                "score": melhor_score,
            }
            resultados.append(resultado)

            if chave_match and email_match:
                yield f"data: {json.dumps({'tipo': 'log', 'msg': f'OK [{melhor_score}%] {orig_name} -> {nome_match} -> {email_match}'})}\n\n"
            elif chave_match:
                yield f"data: {json.dumps({'tipo': 'log', 'msg': f'SEM EMAIL {orig_name} -> {nome_match} (sem email cadastrado)'})}\n\n"
            else:
                yield f"data: {json.dumps({'tipo': 'log', 'msg': f'SEM MATCH {orig_name} -> verifique o nome do arquivo'})}\n\n"

        yield f"data: {json.dumps({'tipo': 'resultado', 'dados': resultados})}\n\n"

    return Response(
        stream_with_context(gerar()),
        mimetype="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no", "Connection": "keep-alive"}
    )


@app.route("/envio_boletos/enviar", methods=["POST"])
@login_required
def enviar_boleto_locatario():
    from email.mime.base import MIMEBase
    from email import encoders as email_encoders

    d = request.get_json() or {}
    arquivo_salvo = (d.get("arquivo_salvo") or "").strip()
    email_dest = (d.get("email") or "").strip()
    locatario = (d.get("locatario") or "").strip()
    mes = (d.get("mes") or "").strip()

    if not email_dest:
        return jsonify({"ok": False, "erro": "Email não informado"})
    if not arquivo_salvo:
        return jsonify({"ok": False, "erro": "Arquivo não informado"})

    pdf_path = UPLOAD_DIR / arquivo_salvo
    if not pdf_path.exists():
        return jsonify({"ok": False, "erro": "Arquivo não encontrado no servidor"})

    smtp = ler_smtp()
    if not smtp["resend_key"] and (not smtp["user"] or not smtp["passw"]):
        return jsonify({"ok": False, "erro": "Email não configurado. Configure em 'Configurar Email'."})

    # Lê PDF como base64
    with open(str(pdf_path), "rb") as f:
        pdf_b64 = base64.b64encode(f.read()).decode()

    # Nome do arquivo para o anexo (sem prefixo loc_)
    nome_anexo = arquivo_salvo[4:] if arquivo_salvo.startswith("loc_") else arquivo_salvo

    mes_texto = mes or "referência"
    assunto = f"Boleto de Aluguel — {mes_texto}"
    if locatario:
        assunto += f" — {locatario}"

    logo_path = BASE / "static" / "logo.png"
    logo_tag = ""
    if logo_path.exists():
        with open(logo_path, "rb") as f:
            b64_logo = base64.b64encode(f.read()).decode()
        logo_tag = f'<img src="data:image/png;base64,{b64_logo}" style="max-width:220px;height:auto;display:block;margin-bottom:16px" alt="Funchal Imoveis">'

    html_body = f"""<!DOCTYPE html>
<html lang="pt-BR"><head><meta charset="UTF-8"></head>
<body style="font-family:'Times New Roman',Times,serif;color:#000;background:#fff;max-width:600px;margin:0 auto;padding:24px">
  {logo_tag}
  <p>Boa tarde,</p>
  <p>Segue em anexo o boleto referente ao m&ecirc;s de <strong>{mes_texto}</strong>.</p>
  <br>
  <p>Atenciosamente,</p>
  <p><strong>Financeiro Funchal Im&oacute;veis</strong></p>
  <hr style="margin-top:32px;border:none;border-top:1px solid #ccc">
  <p style="font-size:9pt;color:#888">Funchal Solu&ccedil;&otilde;es Empresariais e Imobili&aacute;rias Ltda. &mdash; CRECI 21.360-J</p>
</body></html>"""

    if smtp["resend_key"]:
        remetente = "Funchal Imoveis <noreply@funchalimoveis.com.br>"
        reply_to = smtp["user"] or None
        data_resend = {
            "from": remetente,
            "to": [email_dest],
            "subject": assunto,
            "html": html_body,
            "attachments": [{"filename": nome_anexo, "content": pdf_b64}]
        }
        if reply_to:
            data_resend["reply_to"] = [reply_to]
        payload = _json.dumps(data_resend).encode("utf-8")
        req = urllib.request.Request(
            "https://api.resend.com/emails",
            data=payload,
            headers={
                "Authorization": f"Bearer {smtp['resend_key']}",
                "Content-Type": "application/json",
                "User-Agent": "python-urllib/sistema-boletos"
            },
            method="POST"
        )
        try:
            with urllib.request.urlopen(req, timeout=30) as resp:
                return jsonify({"ok": resp.status in (200, 201)})
        except urllib.error.HTTPError as e:
            body = e.read().decode("utf-8", errors="ignore")
            return jsonify({"ok": False, "erro": f"HTTP {e.code}: {body[:300]}"})
        except Exception as e:
            return jsonify({"ok": False, "erro": str(e)})

    # SMTP com anexo
    remetente = smtp["from"] or smtp["user"]
    msg = MIMEMultipart()
    msg["Subject"] = assunto
    msg["From"] = remetente
    msg["To"] = email_dest
    msg.attach(MIMEText(html_body, "html", "utf-8"))
    part = MIMEBase("application", "octet-stream")
    with open(str(pdf_path), "rb") as f:
        part.set_payload(f.read())
    email_encoders.encode_base64(part)
    part.add_header("Content-Disposition", f'attachment; filename="{nome_anexo}"')
    msg.attach(part)
    try:
        with _smtp_connect(smtp, timeout=15) as s:
            s.sendmail(remetente, [email_dest], msg.as_string())
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "erro": str(e)})


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    print(f"Iniciando servidor em http://localhost:{port}")
    app.run(debug=False, host="0.0.0.0", port=port)
